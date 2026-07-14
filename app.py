from flask import Flask, render_template, request, redirect, url_for, flash, session
from models import db, User, Product, Category, Client, Order, OrderDetail, ProductMovement, AuditLog, SystemConfig,OrderKitComponent
from models import ProductImportBolts, CategoryImportBolts, ProductMovementImportBolts
from models import ProductMovement
from models import Payment
from models import Category
from datetime import datetime # Importante para la hora
from flask import send_from_directory
import pandas as pd
import html
from sqlalchemy.exc import IntegrityError # Para capturar el error del SKU
from sqlalchemy import or_, func, text, extract
from sqlalchemy import func
from datetime import datetime, timedelta, date
from flask import send_file
from docxtpl import DocxTemplate, RichText # Importar librería de Word
from werkzeug.utils import secure_filename
from models import SystemConfig
from num2words import num2words
from sqlalchemy import text
from sqlalchemy import or_, func, text, extract, String
from flask_migrate import Migrate
from botocore.client import Config
import os
import io
import subprocess
import tempfile
import requests
import re
import pytz
import boto3
from botocore.exceptions import ClientError
from werkzeug.security import generate_password_hash, check_password_hash
from xhtml2pdf import pisa




def hora_peru():
    # Obtiene la hora exacta de Lima, pero le quita la 'etiqueta' de zona horaria (.replace)
    # para que sea 100% compatible con la base de datos (offset-naive)
    return datetime.now(pytz.timezone('America/Lima')).replace(tzinfo=None)
# --- FUNCIÓN AUXILIAR PARA GUARDAR HISTORIAL ---
def registrar_log(accion, icono='bi-info-circle', color='text-primary'):
    if 'user_id' in session:
        nuevo_log = AuditLog(
            user_id=session['user_id'],
            accion=accion,
            icono=icono,
            color=color
        )
        db.session.add(nuevo_log)
        # Nota: No hacemos commit aquí, asumimos que la función principal lo hará

app = Flask(__name__)

# --- CONFIGURACIÓN DE BASE DE DATOS (Inteligente para la Nube) ---
# Intenta obtener la URL de Render/Railway. Si no existe, usa SQLite en tu PC.
database_url = os.getenv('DATABASE_URL', 'sqlite:///importbolts.db')

# Parche obligatorio para servidores en la nube (SQLAlchemy 1.4+)
if database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'tesis_secreta_123'
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024

# --- CONFIGURACIÓN DE AMAZON S3 ---
# Usamos .strip() para limpiar cualquier espacio invisible o salto de línea
AWS_ACCESS_KEY_ID = os.environ.get('AWS_ACCESS_KEY_ID', '').strip()
AWS_SECRET_ACCESS_KEY = os.environ.get('AWS_SECRET_ACCESS_KEY', '').strip()
AWS_REGION = os.environ.get('AWS_REGION', 'us-east-2').strip()
S3_BUCKET_NAME = os.environ.get('S3_BUCKET_NAME', 'anclajesypernosperu-archivos-145292398833-us-east-2-an').strip()

s3_client = boto3.client(
    's3',
    aws_access_key_id=AWS_ACCESS_KEY_ID,
    aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
    region_name=AWS_REGION,
    config=Config(signature_version='s3v4')
)

# Carpeta EXCLUSIVA para Órdenes de Compra locales (Pre-AWS)
app.config['UPLOAD_FOLDER_OC'] = os.path.join(os.getcwd(), 'uploads_oc')
os.makedirs(app.config['UPLOAD_FOLDER_OC'], exist_ok=True)

# Conectar la base de datos a la app
db.init_app(app)

migrate = Migrate(app, db)

# --- RUTAS BÁSICAS (VISTAS) ---

@app.route('/')
def index():
    if 'user_id' not in session: return redirect(url_for('login'))
    
    rol = session.get('role')
    user_id = session.get('user_id')
    hoy = hora_peru().date()
    
    # --- DATOS COMUNES (Alertas de Stock) ---
    UMBRAL_STOCK = 100 
    total_alertas = Product.query.filter(Product.stock_actual < UMBRAL_STOCK).count()
    alertas_muestra = Product.query.filter(Product.stock_actual < UMBRAL_STOCK).limit(5).all()
    
# ======================================================
    # VISTA 1: ADMIN Y ADMINISTRACIÓN (DASHBOARD BI GLOBAL)
    # ======================================================
    if rol in ['admin', 'administracion']:
        # A. KPIs Financieros
        ventas_hoy = db.session.query(func.sum(Order.total)).filter(func.date(Order.fecha) == hoy).scalar() or 0
        ventas_mes = db.session.query(func.sum(Order.total)).filter(extract('year', Order.fecha) == hoy.year, extract('month', Order.fecha) == hoy.month).scalar() or 0
        pedidos_pendientes = Order.query.filter(Order.estado == 'Pendiente').count()
        
        # B. Ranking de Vendedores
        ranking = db.session.query(
            User.username, 
            User.nombre_completo, 
            func.sum(Order.total).label('total_vendido'),
            func.count(Order.id).label('cantidad_ventas')
        ).join(Order, User.id == Order.vendedor_id) \
         .filter(Order.estado != 'Anulado') \
         .group_by(User.id) \
         .order_by(text('total_vendido DESC')) \
         .limit(5).all()
        
        # C. Productos Más Vendidos
        top_productos = db.session.query(
            Product.nombre,
            func.sum(OrderDetail.cantidad).label('total_qty')
        ).join(OrderDetail).group_by(Product.nombre).order_by(text('total_qty DESC')).limit(5).all()
        
        # D. Predicción
        dias_transcurridos = hoy.day
        promedio_diario = ventas_mes / dias_transcurridos if dias_transcurridos > 0 else 0
        prediccion_fin_mes = promedio_diario * 30

        # ---> [NUEVO CÓDIGO] E. Seguimiento de Cotizaciones (Para el Gerente) <---
        # Traemos las últimas 15 cotizaciones que no estén cerradas/anuladas
        cotizaciones_recientes = Order.query.filter(
            Order.estado.notin_(['Entregado', 'Anulado', 'Despachado'])
        ).order_by(Order.fecha.desc()).limit(15).all()

        # Conteo rápido por estados críticos
        alertas_gerencia = {
            'revision_pre_cliente': Order.query.filter_by(estado='Revision Pre-Cliente').count(),
            'aprobacion_final': Order.query.filter_by(estado='Pendiente Aprobacion Final').count(),
            'observados': Order.query.filter_by(estado='Observado').count(),
            'por_verificar_stock': Order.query.filter_by(estado='Por Verificar').count()
        }
        
        return render_template('dashboard_admin.html', 
                               ventas_hoy=ventas_hoy,
                               ventas_mes=ventas_mes,
                               pedidos_pendientes=pedidos_pendientes,
                               ranking=ranking,
                               top_productos=top_productos,
                               prediccion=prediccion_fin_mes,
                               alertas=alertas_muestra,      
                               total_alertas=total_alertas,
                               cotizaciones_recientes=cotizaciones_recientes, # <-- PASAMOS ESTO
                               alertas_gerencia=alertas_gerencia)             # <-- Y ESTO

    # ======================================================
    # VISTA 2: VENDEDOR (MI RENDIMIENTO PERSONAL)
    # ======================================================
    elif rol == 'vendedor':
        mis_ventas_hoy = db.session.query(func.sum(Order.total)).filter(Order.vendedor_id == user_id, func.date(Order.fecha) == hoy).scalar() or 0
        mis_ventas_mes = db.session.query(func.sum(Order.total)).filter(Order.vendedor_id == user_id, extract('year', Order.fecha) == hoy.year, extract('month', Order.fecha) == hoy.month).scalar() or 0
        mis_pendientes = Order.query.filter_by(vendedor_id=user_id, estado='Pendiente').count()
        
        mis_ultimas = Order.query.filter_by(vendedor_id=user_id).order_by(Order.fecha.desc()).limit(5).all()
        
        return render_template('dashboard_vendedor.html', 
                               hoy=mis_ventas_hoy, 
                               mes=mis_ventas_mes, 
                               pendientes=mis_pendientes,
                               ultimas=mis_ultimas)

    # ======================================================
    # VISTA 3: ALMACÉN (LOGÍSTICA OPERATIVA)
    # ======================================================
    else: # Almacen
        por_despachar = Order.query.filter(Order.estado == 'Pendiente').count()
        en_ruta = Order.query.filter(Order.estado == 'Despachado').count()
        entregados_hoy = Order.query.filter(Order.estado == 'Entregado', func.date(Order.fecha) == hoy).count()
        
        prioritarios = Order.query.filter(Order.estado == 'Pendiente').order_by(Order.fecha_entrega.asc()).limit(5).all()
        
        return render_template('dashboard_almacen.html', 
                               por_despachar=por_despachar,
                               en_ruta=en_ruta,
                               entregados=entregados_hoy,
                               prioritarios=prioritarios,
                               alertas=alertas_muestra,
                               total_alertas=total_alertas)


# --- CONSULTA RUC/DNI (VERSIÓN ROBUSTA) ---
# --- RUTA PARA CONSULTA RUC/DNI (CORREGIDA) ---
# --- EN APP.PY ---


@app.route('/api/consulta_documento', methods=['POST'])
def consulta_documento():
    # 1. Seguridad
    if session.get('user_id') is None: return {'status': 'error', 'msg': 'No autorizado'}, 403
    
    # 2. Variables de Entrada
    numero = request.form.get('numero', '').strip()
    force = request.form.get('force') == 'true'
    usuario_actual = session.get('username', 'Sistema')
    
    # 3. Buscar en Base de Datos Local
    cliente_db = Client.query.filter_by(documento=numero).first()
    
    # --- CANDADO DE SEGURIDAD (AHORRO DE DINERO) ---
    if cliente_db and force:
        # Calculamos el tiempo transcurrido
        tiempo_pasado = hora_peru() - cliente_db.last_updated
        total_segundos = tiempo_pasado.total_seconds()
        horas_pasadas = total_segundos / 3600
        
        # INTERVALO RECOMENDADO: 24 HORAS
        if horas_pasadas < 24:
            usuario_anterior = getattr(cliente_db, 'updated_by', 'Sistema')
            
            if horas_pasadas < 1:
                tiempo_str = f"{int(total_segundos / 60)} minutos"
            else:
                tiempo_str = f"{int(horas_pasadas)} horas"

            return {
                'status': 'error', 
                'msg': f'⛔ AHORRO ACTIVO: Este cliente ya fue actualizado hace {tiempo_str} por {usuario_anterior}. Datos vigentes.'
            }
    # -----------------------------------------------

    # CASO A: Usar dato local (Gratis)
    if cliente_db and not force:
        print(f">>> [AHORRO] Cliente {numero} encontrado en BD Local.")
        return {
            'status': 'success',
            'origen': 'BD',
            'razon_social': cliente_db.nombre,
            'direccion': cliente_db.direccion,
            'telefono': cliente_db.telefono,
            'estado': cliente_db.estado,
            'condicion': cliente_db.condicion,
            
            # DATOS DE AUDITORÍA
            'last_updated': cliente_db.last_updated.strftime('%d/%m %H:%M'),
            'updated_by': getattr(cliente_db, 'updated_by', 'Sistema')
        }

    # CASO B: Consultar API (Costo)
    print(f">>> [API] Consultando datos externos para {numero}...")
    TOKEN = os.getenv('SUNAT_API_KEY')
    # CASO B: Consultar API (Costo)
    print(f">>> [API] Consultando datos externos para {numero}...")
    
    # --- CAMBIO AQUÍ PARA USO LOCAL ---
    TOKEN = "sk_12670.mczJWCBkAFXbV3pYZdD6EoxkwZ7SZSME"
    # ----------------------------------
    
    URL_RUC = "https://api.decolecta.com/v1/sunat/ruc"
    URL_DNI = "https://api.decolecta.com/v1/reniec/dni"
    
    try:
        url = ""
        if len(numero) == 8: url = f"{URL_DNI}?numero={numero}"
        elif len(numero) == 11: url = f"{URL_RUC}?numero={numero}"
        else: return {'status': 'error', 'msg': 'Longitud incorrecta'}

        response = requests.get(url, headers={'Authorization': f'Bearer {TOKEN}'}, timeout=5)
        data = response.json()
        
        if response.status_code == 200: 
            razon = ""
            direccion = ""
            estado = "ACTIVO"
            condicion = "HABIDO"

            if len(numero) == 8: # DNI
                if 'nombres' in data:
                    raw_name = f"{data.get('nombres')} {data.get('apellidoPaterno')} {data.get('apellidoMaterno')}"
                    razon = html.unescape(raw_name) # Limpieza DNI
                    direccion = "-" 
            else: # RUC
                # Obtenemos dato crudo
                raw_razon = data.get('razon_social') or data.get('razonSocial') or data.get('nombre') or ''
                # LIMPIEZA AQUÍ (Esto arregla el &amp;)
                razon = html.unescape(raw_razon)
                
                raw_dir = data.get('direccion', '')
                direccion = html.unescape(raw_dir)
                
                estado = data.get('estado', 'ACTIVO')
                condicion = data.get('condicion', 'HABIDO')

            # 3. GUARDAR O ACTUALIZAR EN BD
            if not cliente_db:
                cliente_db = Client(
                    documento=numero, nombre=razon, direccion=direccion,
                    estado=estado, condicion=condicion,
                    last_updated=hora_peru(),
                    updated_by=usuario_actual
                )
                db.session.add(cliente_db)
            else:
                cliente_db.nombre = razon
                cliente_db.direccion = direccion
                cliente_db.estado = estado
                cliente_db.condicion = condicion
                cliente_db.last_updated = hora_peru()
                cliente_db.updated_by = usuario_actual
            
            db.session.commit()

            return {
                'status': 'success',
                'origen': 'API',
                'razon_social': razon,
                'direccion': direccion,
                'estado': estado,
                'condicion': condicion,
                'last_updated': hora_peru().strftime('%d/%m %H:%M'),
                'updated_by': usuario_actual
            }
        else:
            return {'status': 'error', 'msg': data.get('message', 'No encontrado en SUNAT')}

    except Exception as e:
        print("ERROR API:", str(e))
        return {'status': 'error', 'msg': 'Error de conexión externa'}
        
@app.route('/api/actualizar_telefono_cliente', methods=['POST'])
def actualizar_telefono_cliente():
    if session.get('user_id') is None: return {'status': 'error'}, 403
    
    doc = request.form.get('documento')
    tel = request.form.get('telefono')
    
    cliente = Client.query.filter_by(documento=doc).first()
    if cliente:
        cliente.telefono = tel
        db.session.commit()
        return {'status': 'success'}
    
    return {'status': 'error', 'msg': 'Cliente no encontrado en BD'}

# --- AGREGAR ESTA FUNCIÓN EN APP.PY (Cerca de las otras APIs) ---
# --- EN APP.PY ---

# --- FUNCIÓN INTELIGENTE (CACHÉ) ---
# --- EN APP.PY ---

# --- EN APP.PY ---

def obtener_tipo_cambio(usuario_solicitante="Sistema", forzar=False):
    # 1. Intentar buscar en BD
    config = None
    try:
        config = SystemConfig.query.get('tipo_cambio')
    except:
        pass

    # Usamos la hora del sistema (Tu PC) para evitar líos de zona horaria
    ahora = hora_peru() 
    hoy = ahora.date()

    # DEFINIR HORA DE CORTE SUNAT: 08:30 AM de hoy
    corte_sunat = ahora.replace(hour=8, minute=30, second=0, microsecond=0)

    # --- ANÁLISIS DE DECISIÓN ---
    debo_consultar = False
    motivo_consulta = ""

    if not config:
        debo_consultar = True
        motivo_consulta = "Base de datos vacía"
    
    elif config.updated_at.date() != hoy:
        debo_consultar = True
        motivo_consulta = "Dato es de ayer"
    
    elif forzar:
        # ENFRIAMIENTO (Cool-down)
        # Calculamos segundos pasados desde la última vez
        tiempo_pasado = (ahora - config.updated_at).total_seconds()
        
        # REGLA: Si pasaron menos de 15 min (900 seg), BLOQUEAR para ahorrar
        if tiempo_pasado < 900 and tiempo_pasado >= 0:
            print(f">>> [ESCUDO] Intento rápido ({int(tiempo_pasado)} seg). Usando caché.")
            return float(config.value)
        
        debo_consultar = True
        motivo_consulta = f"Forzado por usuario (Pasaron {int(tiempo_pasado/60)} min)"

    else:
        # Dato de hoy, sin forzar. ¿Es de madrugada?
        # Si el dato guardado es de ANTES de las 8:30 AM y AHORA ya pasó esa hora...
        if config.updated_at < corte_sunat and ahora >= corte_sunat:
            debo_consultar = True
            motivo_consulta = "Actualización automática (Regla 8:30 AM)"
        else:
            return float(config.value)

    # 2. CONSULTA API (Solo si pasó los filtros)
    # 2. CONSULTA API (Solo si pasó los filtros)
    if debo_consultar:
        print(f"--- 🟢 API SUNAT ({motivo_consulta}) ---")
        
        # --- CAMBIO AQUÍ PARA USO LOCAL ---
        TOKEN = "sk_12670.mczJWCBkAFXbV3pYZdD6EoxkwZ7SZSME"
        # ----------------------------------
        
        URL = "https://api.decolecta.com/v1/tipo-cambio/sunat"
        
        try:
            response = requests.get(URL, headers={'Authorization': f'Bearer {TOKEN}'}, timeout=5)
            data = response.json()
            
            if response.status_code == 200:
                precio = float(data.get('venta') or data.get('precio_venta') or data.get('sell_price') or 0.0)
                
                if precio > 0:
                    # 3. GUARDADO EXITOSO
                    if not config:
                        config = SystemConfig(key='tipo_cambio', value=str(precio), updated_at=ahora, updated_by=usuario_solicitante)
                        db.session.add(config)
                    else:
                        config.value = str(precio)
                        config.updated_at = ahora
                        config.updated_by = usuario_solicitante
                    
                    db.session.commit()
                    print(f"--- ✅ BD ACTUALIZADA: {precio} (Por: {usuario_solicitante}) ---")
                    return precio
            
            print(f"--- ⚠️ API ERROR O SIN PRECIO: {data} ---")
            # Si falla, intenta devolver lo que había antes en vez de 3.85
            return float(config.value) if config else 3.850

        except Exception as e:
            print(f"--- ❌ ERROR CONEXIÓN API: {e} ---")
            return float(config.value) if config else 3.850
            
    return 3.850
# --- RUTA API: ACTUALIZAR DÓLAR ---
# --- RUTA API ACTUALIZADA ---
@app.route('/api/tipo_cambio_actual')
def api_tc_actual():
    force_param = request.args.get('force', 'false')
    forzar = (force_param == 'true')
    
    usuario = session.get('username', 'Anonimo')
    
    # Llamada a la lógica
    tc = obtener_tipo_cambio(usuario_solicitante=usuario, forzar=forzar)
    
    # Recuperamos info para mostrar
    config = SystemConfig.query.get('tipo_cambio')
    
    fecha_str = "N/A"
    updated_by = "Sistema"
    es_de_hoy = False
    
    if config:
        fecha_str = config.updated_at.strftime('%d/%m %H:%M')
        updated_by = config.updated_by
        es_de_hoy = (config.updated_at.date() == date.today())

    return {
        'tc': tc, 
        'updated_at': fecha_str,
        'updated_by': updated_by,
        'es_de_hoy': es_de_hoy
    }
    
# 2. NUEVA RUTA: DESCARGAR REPORTE PREDICCIONES (EXCEL)
@app.route('/descargar_reporte_excel')
def descargar_reporte_excel():
    if session.get('role') not in ['admin', 'administracion']: return "Acceso denegado", 403
    
    # Replicamos la lógica de predicción
    productos_activos = db.session.query(
        Product.sku,
        Product.nombre, 
        Product.stock_actual,
        func.sum(ProductMovement.cantidad).label('total_vendido')
    ).join(ProductMovement).filter(
        ProductMovement.tipo == 'SALIDA',
        ProductMovement.fecha >= hora_peru() - timedelta(days=90)
    ).group_by(Product.id).all()
    
    data_excel = []
    
    for p in productos_activos:
        promedio_mensual = p.total_vendido / 3
        prediccion = promedio_mensual * 1.10
        
        estado = "OK"
        faltante = 0
        if prediccion > p.stock_actual:
            estado = "QUIEBRE DE STOCK"
            faltante = prediccion - p.stock_actual
            
        data_excel.append({
            'SKU': p.sku,
            'Producto': p.nombre,
            'Stock Actual': p.stock_actual,
            'Ventas 90 Días': p.total_vendido,
            'Velocidad (Mes)': round(promedio_mensual, 1),
            'Predicción Demanda': round(prediccion, 0),
            'Estado': estado,
            'Sugerencia Compra': round(faltante, 0) if faltante > 0 else 0
        })
    
    # Crear DataFrame y Excel
    df = pd.DataFrame(data_excel)
    
    # Guardar en memoria o temporal
    path = os.path.join(app.config['UPLOAD_FOLDER'], 'Reporte_Predicciones_BI.xlsx')
    df.to_excel(path, index=False)
    
    return send_file(path, as_attachment=True)

# --- API: OBTENER CALIDADES POR FAMILIA (Para el filtro dinámico) ---
@app.route('/api/calidades_de_familia', methods=['POST'])
def calidades_de_familia():
    if session.get('user_id') is None: return {'status': 'error'}, 403
    
    familia = request.form.get('familia')
    
    try:
        # Buscamos en la tabla Productos todas las calidades distintas de esa familia
        # SELECT DISTINCT calidad FROM product WHERE categoria = 'familia' ORDER BY calidad
        calidades = db.session.query(Product.calidad)\
            .filter_by(categoria=familia)\
            .distinct()\
            .order_by(Product.calidad)\
            .all()
        
        # Convertimos el resultado [(CalidadA,), (CalidadB,)] a lista simple ['CalidadA', 'CalidadB']
        lista_calidades = [c[0] for c in calidades if c[0]]
        
        return {'status': 'success', 'calidades': lista_calidades}
        
    except Exception as e:
        return {'status': 'error', 'msg': str(e)}
    
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password, password):
            session['user_id'] = user.id
            session['role'] = user.role
            session['username'] = user.username
            session['nombre'] = user.nombre_completo
            return redirect(url_for('index'))
        else:
            flash('Usuario o contraseña incorrectos')
            
    return render_template('login.html') # Crearemos esto luego
# 1. ACTUALIZAR CONTEXT PROCESSOR (Para la campana inteligente)
@app.context_processor
def inject_notifications():
    if 'user_id' not in session: return dict(alertas_stock=0, historial=[])
    
    # AHORA ES DINÁMICO: Compara stock_actual vs stock_minimo de cada producto
    try:
        count_stock_bajo = Product.query.filter(Product.stock_actual <= Product.stock_minimo).count()
    except:
        count_stock_bajo = 0

    try:
        historial = AuditLog.query.order_by(AuditLog.fecha.desc()).limit(6).all()
    except:
        historial = []
        
    return dict(alertas_stock=count_stock_bajo, historial=historial)

# 2. NUEVA RUTA: EXPORTAR A EXCEL
# --- NUEVA RUTA: EXPORTAR A EXCEL (OPTIMIZADA PARA BAJO CONSUMO DE RAM) ---
@app.route('/producto/exportar')
def exportar_excel():
    import gc # Importamos el recolector de basura
    
    if session.get('role') not in ['admin', 'almacen', 'administracion']: return "No autorizado", 403
    
    # 1. OPTIMIZACIÓN EXTREMA: En lugar de cargar Objetos pesados (Product.query.all()),
    # pedimos solo las columnas exactas (Tuplas ligeras). Esto reduce el uso de RAM un 90%.
    productos = db.session.query(
        Product.sku, Product.nombre, Product.categoria, Product.calidad, 
        Product.ubicacion, Product.stock_actual, Product.stock_minimo, 
        Product.precio_unidad, Product.precio_caja
    ).all()
    
    # Creamos lista de diccionarios
    data = []
    for p in productos:
        data.append({
            'CÓDIGO': p.sku,
            'DESCRIPCIÓN': p.nombre,
            'FAMILIA': p.categoria,
            'CALIDAD': p.calidad,
            'UBICACION': p.ubicacion,
            'STOCK ACTUAL': p.stock_actual,
            'STOCK MÍNIMO': p.stock_minimo,
            'PRECIO UNIT': p.precio_unidad,
            'PRECIO CAJA': p.precio_caja
        })
    
    # Liberamos la memoria RAM de SQLAlchemy antes de procesar el Excel
    del productos
    db.session.expunge_all()
    gc.collect()
    
    # 2. Crear DataFrame y Excel en memoria
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Inventario')
        
        # Ajustar ancho de columnas
        worksheet = writer.sheets['Inventario']
        for idx, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
            worksheet.set_column(idx, idx, max_len)

    output.seek(0)
    
    # 3. LIMPIEZA FINAL: Destruimos Pandas de la memoria RAM
    del data
    del df
    gc.collect()
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Inventario_ImportBolts_{hora_peru().strftime("%Y%m%d")}.xlsx'
    )

# --- NUEVA RUTA: DESCARGAR PLANTILLA VACÍA ---
@app.route('/producto/plantilla')
def descargar_plantilla():
    if session.get('user_id') is None: return redirect(url_for('login'))
    
    # Definir las columnas exactas que el sistema espera
    columnas = [
        'CÓDIGO',       # SKU
        'DESCRIPCIÓN',  # Nombre
        'FAMILIA',      # Categoría
        'CALIDAD',      # Calidad
        'UBICACION',    # Ubicación (Opcional)
        'CANT. ACT.',   # Stock Actual
        'STOCK MÍNIMO'  # Stock Mínimo (Opcional, default 10)
    ]
    
    # Crear un DataFrame vacío con esas columnas
    df = pd.DataFrame(columns=columnas)
    
    # Generar el Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Plantilla_Carga')
        
        # Ajustar ancho de columnas y agregar comentario de ayuda
        workbook = writer.book
        worksheet = writer.sheets['Plantilla_Carga']
        formato_header = workbook.add_format({'bold': True, 'bg_color': '#D7E4BC', 'border': 1})
        
        for idx, col in enumerate(columnas):
            worksheet.set_column(idx, idx, 20)
            worksheet.write(0, idx, col, formato_header)
            
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Plantilla_Importacion_ImportBolts.xlsx'
    )

@app.route('/api/crear_servicio_rapido', methods=['POST'])
def crear_servicio_rapido():
    if session.get('user_id') is None: return {'status': 'error', 'msg': 'No autorizado'}, 403
    
    nombre_servicio = request.form.get('nombre', '').strip().upper()
    if not nombre_servicio: return {'status': 'error', 'msg': 'Nombre vacío'}

    # 1. LIMPIEZA DEL NOMBRE PARA EL SKU
    # Quitamos palabras comunes para que el código sea significativo
    # "SERVICIO DE LAVADO" -> "LAVADO"
    # "SERVICIO TÉCNICO DE MANTENIMIENTO" -> "TÉCNICO DE MANTENIMIENTO"
    palabras_basura = ['SERVICIO DE ', 'SERVICIO ', 'MANO DE OBRA ']
    nombre_limpio = nombre_servicio
    for p in palabras_basura:
        nombre_limpio = nombre_limpio.replace(p, '')
    
    # 2. GENERAR BASE DEL SKU (4 LETRAS)
    # Filtramos solo letras y números, quitamos espacios
    # Ej: "LAVADO" -> "LAVA"
    # Ej: "TORNO CNC" -> "TORN"
    base_limpia = "".join(filter(str.isalnum, nombre_limpio))
    
    if len(base_limpia) < 3:
        # Si el nombre es muy corto (Ej: "A1"), usamos un genérico random para evitar error
        sku_base = "SRV-NUEV"
    else:
        sku_base = f"SRV-{base_limpia[:4]}" # Tomamos las primeras 4

    # 3. VERIFICACIÓN DE DUPLICADOS (Loop de colisión)
    sku_final = sku_base
    contador = 1
    
    # Mientras exista un producto con ese SKU, le sumamos 1
    # Ej: SRV-LAVA, luego SRV-LAVA1, luego SRV-LAVA2...
    while Product.query.filter_by(sku=sku_final).first():
        sku_final = f"{sku_base}{contador}"
        contador += 1

    try:
        # 4. GUARDADO EN BASE DE DATOS
        nuevo_srv = Product(
            sku=sku_final,       # AQUÍ GUARDAMOS EL CÓDIGO ÚNICO
            nombre=nombre_servicio,
            categoria="SERVICIOS",
            calidad="ESTANDAR",
            stock_actual=9999,
            stock_minimo=0,
            precio_unidad=0.0
        )
        db.session.add(nuevo_srv)
        
        # Guardamos log para saber quién creó este servicio nuevo
        registrar_log(f"Creó nuevo servicio: {nombre_servicio} ({sku_final})", "bi-magic", "text-purple")
        
        db.session.commit()
        
        return {
            'status': 'success', 
            'sku': sku_final, 
            'nombre': nombre_servicio,
            'msg': f'Servicio creado correctamente con código: {sku_final}'
        }
        
    except Exception as e:
        db.session.rollback()
        return {'status': 'error', 'msg': str(e)}

# API PARA CARGAR LA LISTA DE SERVICIOS (Para llenar el Select)
@app.route('/api/listar_servicios_activos')
def listar_servicios_activos():
    # Busca todo lo que empiece con SKU 'SRV-' o categoría 'SERVICIOS'
    servicios = Product.query.filter(
        or_(Product.sku.like('SRV-%'), Product.categoria == 'SERVICIOS')
    ).order_by(Product.nombre).all()
    
    lista = [{'sku': s.sku, 'nombre': s.nombre} for s in servicios]
    return {'servicios': lista}

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# 1. ACTUALIZAR RUTA INVENTARIO (Para ver categorías nuevas vacías)
# --- RUTA INVENTARIO (CORREGIDA) ---
# --- RUTA INVENTARIO (CORREGIDA CON REGISTRO GLOBAL) ---
@app.route('/inventario')
def inventario():
    if session.get('user_id') is None: return redirect(url_for('login'))
    
    # 1. Parámetros
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('busqueda', '')
    cat_filtro = request.args.get('categoria', 'todos')
    calidad_filtro = request.args.get('calidad', 'todos')
    stock_bajo = request.args.get('stock_bajo') # Recibe 'on' o None

    # 2. Query Base
    query = Product.query

    if search:
        query = query.filter(or_(Product.nombre.ilike(f"%{search}%"), Product.sku.ilike(f"%{search}%")))
    
    if cat_filtro != 'todos':
        query = query.filter(Product.categoria == cat_filtro)

    if calidad_filtro != 'todos':
        query = query.filter(Product.calidad == calidad_filtro)

    # --- CORRECCIÓN CRÍTICA ---
    if stock_bajo == 'on':
        query = query.filter(Product.stock_actual <= Product.stock_minimo)

    # Listas para los selects
    cats_db = Category.query.order_by(Category.nombre).all()
    lista_categorias = [c.nombre for c in cats_db]
    calidades = db.session.query(Product.calidad).distinct().order_by(Product.calidad).all()
    lista_calidades = [c[0] for c in calidades if c[0]]

    # --- NUEVO: OBTENER ESTADOS ÚNICOS PARA EL AUTOCOMPLETADO ---
    estados = db.session.query(Product.estado).filter(Product.estado != None, Product.estado != '').distinct().order_by(Product.estado).all()
    lista_estados = [e[0] for e in estados]

    # Ordenar y Paginar
    if stock_bajo == 'on':
        query = query.order_by(Product.stock_actual.asc())
    else:
        query = query.order_by(Product.id.asc())

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    productos = pagination.items
    
    # --- NUEVO: CONSULTA DE ÚLTIMA IMPORTACIÓN MASIVA ---
    info_importacion = SystemConfig.query.get('ultima_importacion')
    
    return render_template('inventario.html', 
                           productos=productos, 
                           lista_categorias=lista_categorias, 
                           lista_calidades=lista_calidades,
                           pagination=pagination,
                           search=search,
                           cat_filtro=cat_filtro,
                           calidad_filtro=calidad_filtro,
                           stock_bajo=stock_bajo, 
                           limit=per_page,
                           lista_estados=lista_estados,
                           info_importacion=info_importacion) # <-- Pasado al HTML
                           

# --- API: OBTENER SIGUIENTE SKU (Magia Automática) ---
@app.route('/api/next_sku/<int:category_id>')
def get_next_sku(category_id):
    cat = Category.query.get_or_404(category_id)
    siguiente_num = cat.contador + 1
    # Genera formato: PER-005 (Rellena con ceros hasta 3 dígitos)
    sku_sugerido = f"{cat.prefijo}-{str(siguiente_num).zfill(3)}"
    return {'sku': sku_sugerido, 'prefijo': cat.prefijo}

# --- 1. API PARA BUSCAR CLIENTE (NUEVO) ---
# Pégalo junto a las otras rutas de API
@app.route('/api/cliente/<documento>')
def buscar_cliente(documento):
    cliente = Client.query.filter_by(documento=documento).first()
    if cliente:
        return {
            'encontrado': True,
            'nombre': cliente.nombre,
            'telefono': cliente.telefono,
            'direccion': cliente.direccion
        }
    return {'encontrado': False}

# Modifica esta función en app.py para aceptar el ID 0 como "todos"

@app.route('/api/productos_por_categoria/<int:category_id>')
def get_productos_por_categoria(category_id):
    try:
        if category_id == 0:
            productos = Product.query.limit(500).all()
        else:
            cat = Category.query.get_or_404(category_id)
            productos = Product.query.filter(Product.categoria == cat.nombre).all()
        
        lista = []
        for p in productos:
            # Mandamos el nombre limpio y el estado como un dato independiente
            lista.append({
                'id': p.id,
                'sku': p.sku,
                'nombre': p.nombre, 
                'stock': p.stock_actual,
                'p_unidad': p.precio_unidad,
                'calidad': p.calidad,
                'estado': p.estado if p.estado else '' # <-- ESTADO SEPARADO
            })
        return {'productos': lista}
        
    except Exception as e:
        print(f"API Error: {e}")
        return {'productos': []}

# --- GESTIÓN DE USUARIOS (ADMIN) ---

@app.route('/usuarios')
def gestion_usuarios():
    # Seguridad: Solo admin
    if session.get('role') != 'admin': 
        return "Acceso denegado", 403
    
    usuarios = User.query.all()
    return render_template('usuarios.html', usuarios=usuarios)

# --- EN APP.PY ---

@app.route('/usuarios/guardar', methods=['POST'])
def guardar_usuario():
    if session.get('role') != 'admin': return "Acceso denegado", 403
    
    # 1. CAPTURA DE DATOS
    user_id = request.form.get('user_id')
    
    # Datos personales
    nombres = request.form.get('nombres', '').strip().title()
    apellidos = request.form.get('apellidos', '').strip().title()
    
    # Si estamos editando y usan el formulario viejo, recuperamos nombre_completo directo
    nombre_completo_form = request.form.get('nombre_completo', '').strip()
    
    # Lógica: Si hay nombres separados, los unimos. Si no, usamos el completo.
    if nombres and apellidos:
        nombre_final = f"{nombres} {apellidos}"
    else:
        nombre_final = nombre_completo_form

    # Credenciales y Contacto
    username = request.form['username'].strip().lower() # Siempre minúsculas
    password = request.form['password']
    rol = request.form['role']
    celular = request.form.get('celular', '').strip()
    cargo = request.form.get('cargo_formal', '').strip().upper()
    email = request.form.get('email_empresa', '').strip().lower()
    
    try:
        # 2. VALIDACIONES DE NEGOCIO (BACKEND)
        
        # A. Validación de Celular Perú (9 dígitos, empieza con 9)
        if celular:
            if not celular.isdigit() or len(celular) != 9 or not celular.startswith('9'):
                flash('⛔ Error: El celular debe tener 9 dígitos y empezar con 9.', 'error')
                return redirect(url_for('gestion_usuarios'))

        # B. Validación de Duplicados (Username)
        # Buscamos si existe alguien con ese usuario, PERO que no sea el mismo que estamos editando
        usuario_existente = User.query.filter_by(username=username).first()
        if usuario_existente:
            # Si es nuevo (no hay user_id) O si es edición pero el ID es diferente
            if not user_id or (user_id and usuario_existente.id != int(user_id)):
                flash(f'⛔ Error: El usuario "{username}" ya existe. Elija otro.', 'error')
                return redirect(url_for('gestion_usuarios'))

        # 3. GUARDADO / ACTUALIZACIÓN
        if user_id:
            # --- EDICIÓN ---
            usuario = User.query.get_or_404(user_id)
            usuario.username = username
            usuario.nombre_completo = nombre_final
            usuario.role = rol
            usuario.celular = celular
            usuario.cargo_formal = cargo
            usuario.email_empresa = email
            
            if password:
                usuario.password = generate_password_hash(password)
                flash(f'✅ Perfil de {nombres} actualizado con nueva contraseña.')
            else:
                flash(f'✅ Perfil de {nombres} actualizado.')
        else:
            # --- CREACIÓN ---
            if not password:
                flash('⛔ Error: La contraseña es obligatoria para nuevos usuarios.', 'error')
                return redirect(url_for('gestion_usuarios'))
                
            nuevo = User(
                username=username, 
                nombre_completo=nombre_final, 
                password=generate_password_hash(password), 
                role=rol, 
                celular=celular,
                cargo_formal=cargo, 
                email_empresa=email
            )
            db.session.add(nuevo)
            flash(f'✅ Usuario "{username}" creado exitosamente.')
            
        db.session.commit()

    except Exception as e:
        db.session.rollback()
        flash(f'Error crítico: {str(e)}', 'error')
        
    return redirect(url_for('gestion_usuarios'))


# --- EN APP.PY ---

@app.route('/perfil', methods=['GET', 'POST'])
def perfil_usuario():
    if 'user_id' not in session: return redirect(url_for('login'))
    
    usuario = User.query.get_or_404(session['user_id'])
    
    if request.method == 'POST':
        try:
            # 1. ACTUALIZAR DATOS PERSONALES
            # El usuario no puede cambiar su Login ni su Rol, solo datos de contacto
            usuario.nombre_completo = request.form['nombre_completo'].strip().title()
            usuario.celular = request.form.get('celular', '').strip()
            usuario.email_empresa = request.form.get('email_empresa', '').strip().lower()
            
            # 2. CAMBIO DE CONTRASEÑA (Lógica Segura)
            pass_actual = request.form.get('current_password')
            pass_nueva = request.form.get('new_password')
            pass_confirm = request.form.get('confirm_password')
            
            if pass_nueva: # Si intentó escribir una nueva clave
                if not pass_actual:
                    flash('⛔ Para cambiar la contraseña, debe ingresar su contraseña actual.', 'error')
                    return redirect(url_for('perfil_usuario'))
                
                if not check_password_hash(usuario.password, pass_actual):
                    flash('⛔ La contraseña actual ingresada es incorrecta.', 'error')
                    return redirect(url_for('perfil_usuario'))  
                    
                if pass_nueva != pass_confirm:
                    flash('⛔ Las nuevas contraseñas no coinciden.', 'error')
                    return redirect(url_for('perfil_usuario'))
                
                # Si todo ok, cambiamos la clave
                usuario.password = generate_password_hash(pass_nueva)
                flash('✅ Contraseña actualizada correctamente.', 'success')
            
            db.session.commit()
            flash('✅ Datos de perfil actualizados.', 'success')
            return redirect(url_for('perfil_usuario'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar: {str(e)}', 'error')

    return render_template('perfil.html', u=usuario)

@app.route('/usuarios/eliminar/<int:user_id>')
def eliminar_usuario(user_id):
    if session.get('role') != 'admin': return "Acceso denegado", 403
    
    # Protección: No te puedes borrar a ti mismo
    if user_id == session.get('user_id'):
        flash('Error: No puedes eliminar tu propia cuenta mientras estás conectado.')
        return redirect(url_for('gestion_usuarios'))
    
    usuario = User.query.get_or_404(user_id)
    db.session.delete(usuario)
    db.session.commit()
    flash('Usuario eliminado permanentemente.')
    
    return redirect(url_for('gestion_usuarios'))

# --- MODIFICAR LA RUTA NUEVA_VENTA EN APP.PY ---
# --- MODIFICAR LA RUTA NUEVA_VENTA EN APP.PY ---
@app.route('/nueva_venta', methods=['GET', 'POST'])
def nueva_venta():
    # --- MÉTODO POST (GUARDAR COTIZACIÓN) ---
    if request.method == 'POST':
        try:
            data = request.get_json()

            # >>> IMPRESIÓN DE CONTROL PARA NUEVA COTIZACIÓN <<<
            print("\n========================================")
            print("👉 INTENTANDO CREAR NUEVA COTIZACIÓN:")
            print(f"Agencia: {data.get('agencia')} | Calidad: {data.get('control_calidad')} | Penalidad: {data.get('penalidad')}")
            print("========================================\n")
            
            # 1. GESTIÓN DEL CLIENTE
            cliente = Client.query.filter_by(documento=data.get('cliente_ruc')).first()
            if not cliente:
                cliente = Client(
                    documento=data.get('cliente_ruc'),
                    nombre=data.get('cliente_nombre'),
                    telefono=data.get('cliente_tel'),
                    direccion=data.get('cliente_dir'),
                    estado='ACTIVO', condicion='HABIDO', last_updated=hora_peru()
                )
                db.session.add(cliente)
            else:
                # Actualizar datos si cambiaron
                cliente.nombre = data.get('cliente_nombre')
                cliente.direccion = data.get('cliente_dir')
                cliente.telefono = data.get('cliente_tel')
            
            db.session.flush() # Para asegurar que tenemos el ID del cliente

            # Procesar fecha de entrega
            fecha_str = data.get('fecha_entrega')
            fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date() if fecha_str else None

            # Procesar fecha de vencimiento (Validez Oferta)
            dias_validez = 5 # Default
            validez_txt = data.get('validez_oferta', '5 días')
            try:
                # Extrae solo los números del texto "15 días" -> 15
                import re
                nums = re.findall(r'\d+', validez_txt)
                if nums: dias_validez = int(nums[0])
            except:
                pass
            fecha_vencimiento_calc = hora_peru().date() + timedelta(days=dias_validez)

            tipo_entrega = data['tipo_entrega']
            dir_entrega_final = data['direccion_entrega']

            if tipo_entrega == 'Recojo':
                dir_entrega_final = "RECOJO EN TIENDA / ALMACÉN"
            
            # ==============================================================================
            # 2. VALIDACIÓN Y CÁLCULO SEGURO EN BACKEND (SEGURIDAD FINANCIERA)
            # ==============================================================================
            # No confiamos en data['total'] del frontend. Lo recalculamos aquí.
            
            subtotal_calculado = 0.0
            
            # Primer barrido para calcular totales y validar costos
            for item in data['items']:
                cantidad = float(item.get('cantidad', 0))
                precio_unit = float(item.get('precio', 0))
                
                # --- VALIDACIÓN DE MARGEN (Opcional: Descomentar raise para bloquear) ---
                if item.get('tipo') == 'PRODUCTO':
                    prod_db = Product.query.get(item.get('id'))
                    if prod_db:
                        # Convertir a USD base para comparar con costo
                        precio_en_usd = precio_unit
                        if data.get('moneda') == 'PEN':
                            tc_val = float(data.get('tc', 1))
                            if tc_val > 0: precio_en_usd = precio_unit / tc_val
                        
                        if precio_en_usd < prod_db.costo_referencial:
                            print(f"ALERTA: Producto {prod_db.sku} vendido bajo costo.")
                            # raise Exception(f"El producto {prod_db.sku} tiene un precio menor al costo.") 

                linea_total = round(cantidad * precio_unit, 2)
                subtotal_calculado += linea_total

            # Recalcular Descuentos y Totales Matemáticamente
            tipo_desc = data.get('descuento_tipo')
            val_desc = float(data.get('descuento_valor', 0))
            monto_descuento = 0.0

            if tipo_desc == 'PORCENTAJE':
                monto_descuento = round(subtotal_calculado * (val_desc / 100), 2)
            else:
                monto_descuento = val_desc
            
            # Evitar descuentos negativos o mayores al total
            if monto_descuento < 0: monto_descuento = 0
            if monto_descuento > subtotal_calculado: monto_descuento = subtotal_calculado

            # Totales Finales Oficiales
            subtotal_neto_final = subtotal_calculado - monto_descuento
            igv_final = round(subtotal_neto_final * 0.18, 2)
            total_final = round(subtotal_neto_final + igv_final, 2)

            # 3. CREAR LA ORDEN (CABECERA) CON VALORES SEGUROS
            nueva_orden = Order(
                cliente_id=cliente.id, 
                vendedor_id=session['user_id'], 
                fecha=hora_peru(),
                
                # USAMOS LOS VALORES CALCULADOS EN BACKEND
                subtotal=subtotal_neto_final,
                igv=igv_final,
                total=total_final,
                descuento_tipo=tipo_desc,
                descuento_valor=val_desc,
                descuento_total=monto_descuento,

                moneda=data.get('moneda', 'PEN'),            
                tipo_cambio=float(data.get('tc', 1.0)),      
                
                # Datos Logísticos
                tipo_entrega=tipo_entrega,
                direccion_envio=dir_entrega_final,
                fecha_entrega=fecha_obj,
                agencia=data.get('agencia', 'NO REQUIERE'),              # <--- ASEGÚRATE DE ESTO
                control_calidad=data.get('control_calidad', 'NO'),       # <--- ASEGÚRATE DE ESTO
                penalidad=data.get('penalidad', 'NO'),                   # <--- ASEGÚRATE DE ESTO
                # fecha_vencimiento=fecha_vencimiento_calc, # Descomentar si agregaste el campo al modelo

                estado='Cotizacion',
                
                # Campos adicionales
                atencion=data.get('cliente_atte'),
                orden_compra=data.get('orden_compra'),
                
                # Condiciones Comerciales
                condicion_pago=data.get('condicion_pago'),
                validez_oferta=data.get('validez_oferta'),
                plazo_entrega_texto=data.get('plazo_entrega_texto'),
                observacion=data.get('observacion')
            )
            db.session.add(nueva_orden)
            db.session.flush() # Para obtener el ID de la orden
            
            # 4. GUARDAR DETALLES (Items)
            for item in data['items']:
                tipo_item = item.get('tipo', 'PRODUCTO') 
                
                detalle = OrderDetail(
                    order_id=nueva_orden.id,
                    item_type=tipo_item, 
                    cantidad=int(item['cantidad']),
                    precio_aplicado=float(item['precio']),
                    subtotal=float(item['subtotal']), # Guardamos el visual, pero el header ya cuadra
                    tipo_precio_usado=item.get('tipo_precio', 'Manual'),

                    # --- GUARDADO DE DESCUENTOS ---
                    precio_base=float(item.get('precioBase', item['precio'])),
                    desc_tipo=item.get('desc_tipo', ''),
                    desc_valor=float(item.get('desc_valor', 0.0)),
                    desc_label=item.get('desc_label', '')
                )

                # --- VINCULACIÓN DE PRODUCTOS/SERVICIOS ---
                if tipo_item == 'PRODUCTO':
                    detalle.product_id = item['id']
                
                elif tipo_item == 'FABRICACION':
                    # Intentamos buscar el Servicio por su SKU
                    sku_buscado = item.get('sku')
                    if sku_buscado:
                        prod_servicio = Product.query.filter_by(sku=sku_buscado).first()
                        if prod_servicio:
                            detalle.product_id = prod_servicio.id 
                    
                    # Guardamos los textos
                    detalle.nombre_personalizado = item.get('descripcion_glb', item['nombre'])
                    detalle.nombre_personalizado_titulo = item.get('titulo_glb', '')

                else: # GLB
                    detalle.nombre_personalizado = item.get('descripcion_glb', item['nombre'])
                    detalle.nombre_personalizado_titulo = item.get('titulo_glb', '')

                db.session.add(detalle)
                db.session.flush() # Necesitamos el ID del detalle

                # SI ES UN KIT (GLB), GUARDAMOS SUS COMPONENTES
                if tipo_item == 'GLB' and 'componentes' in item and item['componentes']:
                    for comp in item['componentes']:
                        nuevo_comp = OrderKitComponent(
                            order_detail_id=detalle.id,
                            product_id=comp['id'],      
                            cantidad_requerida=int(comp['qty']) 
                        )
                        db.session.add(nuevo_comp)
                
            db.session.commit()
            
            # Formato visual del ID para el usuario (Ej: COT-00052)
            codigo_visual = f"{nueva_orden.id:05d}" 

            return {
                'status': 'success', 
                'order_id': nueva_orden.id,
                'codigo_visual': codigo_visual
            }
            
        except Exception as e:
            db.session.rollback()
            print(f"ERROR AL GUARDAR VENTA: {e}")
            return {'status': 'error', 'msg': str(e)}, 500

    # --- MÉTODO GET (MOSTRAR PANTALLA) ---
# --- MÉTODO GET (MOSTRAR PANTALLA DE NUEVA VENTA) ---
    productos = Product.query.all()
    categorias = Category.query.all()
    
    tc_hoy = obtener_tipo_cambio(usuario_solicitante="Sistema Automático")
    config_tc = SystemConfig.query.get('tipo_cambio')
    
    # --- NUEVO: CONSULTA DE ÚLTIMA IMPORTACIÓN MASIVA ---
    info_importacion = SystemConfig.query.get('ultima_importacion')
    
    return render_template('nueva_venta.html', 
                           productos=productos, 
                           categorias=categorias, 
                           tc=tc_hoy,
                           updated_at=config_tc.updated_at.strftime('%d/%m %H:%M') if config_tc else None,
                           updated_by=config_tc.updated_by if config_tc else None,
                           info_importacion=info_importacion) # <-- Pasado al HTML

# --- EN APP.PY (Función DESCARGAR MAESTRA) ---
from xhtml2pdf import pisa
import io

@app.route('/descargar_cotizacion/<int:order_id>')
def descargar_cotizacion(order_id):
    # 1. Obtener datos básicos
    orden = Order.query.get_or_404(order_id)
    vendedor = orden.vendedor
    
    # --- A. LÓGICA DE PLANTILLA Y MODO (NUEVO) ---
    modo = request.args.get('modo', 'default') # 'almacen', 'valorizado', o 'default'
    es_aprobado = orden.estado in ['Aprobado', 'Despachado', 'Entregado']
    
    # Configuración por defecto
    template_name = "plantilla_cotizacion.docx"
    titulo_doc = "COTIZACIÓN"
    codigo_visual = f"COT-{orden.id:05d}"
    mostrar_precios = True

    if es_aprobado:
        # ES UNA ORDEN DE PEDIDO (OP)
        template_name = "plantilla_orden_pedido.docx"
        titulo_doc = "ORDEN DE PEDIDO"
        codigo_visual = f"OP-{orden.id:05d}"
        
        if modo == 'almacen':
            mostrar_precios = False
            titulo_doc += " (ALMACÉN)"
    
    # Cargar la plantilla seleccionada
    doc = DocxTemplate(template_name)

    # --- B. PROCESAMIENTO DE DATOS (TU LÓGICA ORIGINAL) ---
    cargo_mostrar = vendedor.cargo_formal if vendedor.cargo_formal else "Asesor Comercial"
    email_texto = vendedor.email_empresa if vendedor.email_empresa else "ventas@importbolts.com"
    celular_texto = vendedor.celular if vendedor.celular else ""

    subtotal_bruto = orden.subtotal + orden.descuento_total

    # Fecha de Entrega
    texto_entrega = "Inmediata / A coordinar"
    if orden.fecha_entrega:
        texto_entrega = orden.fecha_entrega.strftime("%d/%m/%Y")
    
    simbolo = "S/" if orden.moneda == 'PEN' else "$"
    nombre_moneda = "SOLES" if orden.moneda == 'PEN' else "DOLARES AMERICANOS"

    # Procesamiento de Items
    lista_items = []
    i = 1
    
    for d in orden.details:
        sku_final = "SERV"
        if d.product:
            sku_final = d.product.sku
        elif d.item_type == 'FABRICACION':
            mapa_skus = {
                'SERVICIO DE CORTE': 'SRV-CORT',
                'SERVICIO DE SOLDADURA': 'SRV-SOLD',
                'SERVICIO DE GALVANIZADO': 'SRV-GALV',
                'SERVICIO DE ZINCADO': 'SRV-ZINC',
                'SERVICIO DE ROSCADO': 'SRV-ROSC',
                'SERVICIO DE TROPICALIZADO': 'SRV-TROP',
                'SERVICIO GENERAL': 'SRV-GEN'
            }
            titulo_limpio = d.nombre_personalizado_titulo.upper() if d.nombre_personalizado_titulo else ""
            sku_final = mapa_skus.get(titulo_limpio, 'SRV-GEN')
            
            if sku_final == 'SRV-GEN':
                prod_db = Product.query.filter_by(nombre=titulo_limpio).first()
                if prod_db: sku_final = prod_db.sku

        elif d.item_type == 'GLB':
            sku_final = "GLB-001" 

        # Lógica RichText (Descripción)
        descripcion_rich = RichText()
        estilo_fuente = {'font': 'Calibri', 'size': 18} # ~9pt

        if d.product and d.item_type == 'PRODUCTO':
            descripcion_rich.add(d.product.nombre, **estilo_fuente)
        else:
            titulo = d.nombre_personalizado_titulo.upper() if d.nombre_personalizado_titulo else ""
            cuerpo = d.nombre_personalizado.upper() if d.nombre_personalizado else "" 
            
            if titulo:
                descripcion_rich.add(titulo, bold=True, **estilo_fuente)
                if cuerpo: descripcion_rich.add(" ", **estilo_fuente) 
            if cuerpo:
                descripcion_rich.add(cuerpo, **estilo_fuente)

        # Unidad de Medida
        unidad_final = "UND" 
        if d.item_type == 'FABRICACION': unidad_final = "SRV"
        elif d.item_type == 'GLB': unidad_final = "GLB"
        elif d.product and hasattr(d.product, 'unidad_medida'): 
            unidad_final = d.product.unidad_medida or "UND"
        
        lista_items.append({
            'item': i,
            'sku': sku_final,
            'cant': d.cantidad,
            'um': unidad_final,
            'desc': descripcion_rich, 
            'unit': f"{d.precio_aplicado:,.2f}", 
            'subtotal': f"{d.subtotal:,.2f}"
        })
        i += 1
    
    # Conversión a letras
    try:
        total_float = float(orden.total)
        parte_entera = int(total_float)
        parte_decimal = int(round((total_float - parte_entera) * 100))
        letras = num2words(parte_entera, lang='es').upper()
        total_letras = f"{letras} CON {parte_decimal:02d}/100 {nombre_moneda}"
    except:
        total_letras = "---"

    # --- 3. CONTEXTO FINAL (Fusionado) ---
    context = {
        'titulo_documento': titulo_doc,
        'codigo_pedido': codigo_visual,
        'mostrar_precios': mostrar_precios,
        'fecha': orden.fecha.strftime("%d/%m/%Y"),
        'cliente_nombre': orden.cliente.nombre,
        'cliente_ruc': orden.cliente.documento,
        'cliente_direccion_fiscal': orden.cliente.direccion,
        'cliente_telefono': orden.cliente.telefono or "",
        'contacto_atte': orden.atencion or "",
        'orden_compra': orden.orden_compra or "",
        'tipo_entrega': orden.tipo_entrega,
        'lugar_entrega': orden.direccion_envio,
        'plazo_entrega': texto_entrega,
        'vendedor_nombre': orden.vendedor.nombre_completo,
        'vendedor_cargo': cargo_mostrar,
        'vendedor_email': email_texto,
        'vendedor_celular': celular_texto,
        'tbl_contents': lista_items,
        'simbolo': simbolo,
        'subtotal_bruto': f"{subtotal_bruto:,.2f}",
        'label_descuento': f"Descuento ({int(orden.descuento_valor)}%)" if orden.descuento_tipo == 'PORCENTAJE' else "Descuento",
        'monto_descuento': f"- {orden.descuento_total:,.2f}",
        'subtotal_neto': f"{orden.subtotal:,.2f}",
        'subtotal': f"{orden.subtotal:,.2f}",
        'igv': f"{orden.igv:,.2f}",
        'total': f"{orden.total:,.2f}",
        'son_letras': total_letras,
        'condicion_pago': orden.condicion_pago or "Contado",
        'validez': orden.validez_oferta or "5 días",
        'observacion': orden.observacion or "",
        'vendedor_usuario': vendedor.username,
        'almacenero_nombre': orden.almacenero_nombre or "VERIFICADO",
        'fecha_almacen': orden.fecha_verificacion_almacen.strftime('%d/%m/%Y %H:%M') if orden.fecha_verificacion_almacen else "---",
        'gerente_nombre': orden.gerente_nombre or "APROBADO",
        'fecha_gerencia': orden.fecha_aprobacion.strftime('%d/%m/%Y %H:%M') if orden.fecha_aprobacion else "---"
    }
    
    doc.render(context)
    
    # Nombre del archivo dinámico
    suffix = "_ALMACEN" if modo == 'almacen' else ""
    nombre_archivo_base = f"{codigo_visual}_{orden.cliente.nombre[:10]}{suffix}"
    
    # Limpiar caracteres raros del nombre de archivo
    nombre_archivo_base = "".join([c for c in nombre_archivo_base if c.isalnum() or c in (' ', '.', '-', '_')]).strip()

    # =====================================================================
    # NUEVA LÓGICA: CONVERSIÓN A PDF USANDO LIBREOFFICE EN MEMORIA TEMPORAL
    # =====================================================================
    with tempfile.TemporaryDirectory() as tmpdir:
        docx_path = os.path.join(tmpdir, f"{nombre_archivo_base}.docx")
        pdf_path = os.path.join(tmpdir, f"{nombre_archivo_base}.pdf")
        
        # 1. Guardar el docx temporalmente
        doc.save(docx_path)
        
        try:
            # 2. Comando para convertir a PDF
            comando = [
                'libreoffice', '--headless', '--convert-to', 'pdf', 
                '--outdir', tmpdir, docx_path
            ]
            
            # (Opcional) Si estás probando en tu computadora con Windows localmente:
            if os.name == 'nt':
                # Ruta por defecto donde se instala LibreOffice en Windows
                ruta_windows = r"C:\Program Files\LibreOffice\program\soffice.exe"
                if os.path.exists(ruta_windows):
                    comando[0] = ruta_windows
                else:
                    print("Advertencia: LibreOffice no encontrado en la ruta por defecto de Windows.")

            # Ejecutar la conversión
            subprocess.run(comando, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            
            # 3. Leer el PDF generado
            with open(pdf_path, 'rb') as f:
                pdf_data = f.read()
                
            # 4. Enviar el PDF al usuario
            return send_file(
                io.BytesIO(pdf_data),
                as_attachment=True,
                download_name=f"{nombre_archivo_base}.pdf",
                mimetype='application/pdf'
            )
            
        except Exception as e:
            print(f"Error al convertir a PDF: {str(e)}")
            # FALLBACK DE SEGURIDAD: Si falla la conversión a PDF, descarga el Word normal
            output = io.BytesIO()
            doc.save(output)
            output.seek(0)
            return send_file(
                output, 
                as_attachment=True, 
                download_name=f"{nombre_archivo_base}.docx",
                mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            )


@app.route('/descargar_cotizacion_v2/<int:order_id>')
def descargar_cotizacion_v2(order_id):
    orden = Order.query.get_or_404(order_id)
    vendedor = orden.vendedor
 
    # --- A. LÓGICA DE PLANTILLA Y MODO (igual que tu ruta original) ---
    modo = request.args.get('modo', 'default')
    es_aprobado = orden.estado in ['Aprobado', 'Despachado', 'Entregado']
 
    titulo_doc = "COTIZACIÓN"
    codigo_visual = f"COT-{orden.id:05d}"
    mostrar_precios = True
 
    if es_aprobado:
        titulo_doc = "ORDEN DE PEDIDO"
        codigo_visual = f"OP-{orden.id:05d}"
        if modo == 'almacen':
            mostrar_precios = False
            titulo_doc += " (ALMACÉN)"
 
    # --- B. PROCESAMIENTO DE DATOS (igual que tu lógica original) ---
    cargo_mostrar = vendedor.cargo_formal if vendedor.cargo_formal else "Asesor Comercial"
    email_texto = vendedor.email_empresa if vendedor.email_empresa else "ventas@importbolts.com"
    celular_texto = vendedor.celular if vendedor.celular else ""
 
    subtotal_bruto = orden.subtotal + orden.descuento_total
 
    texto_entrega = "Inmediata / A coordinar"
    if orden.fecha_entrega:
        texto_entrega = orden.fecha_entrega.strftime("%d/%m/%Y")
 
    simbolo = "S/" if orden.moneda == 'PEN' else "$"
    nombre_moneda = "SOLES" if orden.moneda == 'PEN' else "DOLARES AMERICANOS"
 
    # Procesamiento de Items (sin RichText, ahora HTML simple con <b>)
    lista_items = []
    i = 1
 
    for d in orden.details:
        sku_final = "SERV"
        if d.product:
            sku_final = d.product.sku
        elif d.item_type == 'FABRICACION':
            mapa_skus = {
                'SERVICIO DE CORTE': 'SRV-CORT',
                'SERVICIO DE SOLDADURA': 'SRV-SOLD',
                'SERVICIO DE GALVANIZADO': 'SRV-GALV',
                'SERVICIO DE ZINCADO': 'SRV-ZINC',
                'SERVICIO DE ROSCADO': 'SRV-ROSC',
                'SERVICIO DE TROPICALIZADO': 'SRV-TROP',
                'SERVICIO GENERAL': 'SRV-GEN'
            }
            titulo_limpio = d.nombre_personalizado_titulo.upper() if d.nombre_personalizado_titulo else ""
            sku_final = mapa_skus.get(titulo_limpio, 'SRV-GEN')
 
            if sku_final == 'SRV-GEN':
                prod_db = Product.query.filter_by(nombre=titulo_limpio).first()
                if prod_db:
                    sku_final = prod_db.sku
 
        elif d.item_type == 'GLB':
            sku_final = "GLB-001"
 
        if d.product and d.item_type == 'PRODUCTO':
            descripcion_html = d.product.nombre
        else:
            titulo = d.nombre_personalizado_titulo.upper() if d.nombre_personalizado_titulo else ""
            cuerpo = d.nombre_personalizado.upper() if d.nombre_personalizado else ""
            partes = []
            if titulo:
                partes.append(f"<b>{titulo}</b>")
            if cuerpo:
                partes.append(cuerpo)
            descripcion_html = " ".join(partes)
 
        unidad_final = "UND"
        if d.item_type == 'FABRICACION':
            unidad_final = "SRV"
        elif d.item_type == 'GLB':
            unidad_final = "GLB"
        elif d.product and hasattr(d.product, 'unidad_medida'):
            unidad_final = d.product.unidad_medida or "UND"
 
        lista_items.append({
            'item': i,
            'sku': sku_final,
            'um': unidad_final,
            'desc': descripcion_html,
            'cant': d.cantidad,
            'unit': f"{d.precio_aplicado:,.2f}",
            'subtotal': f"{d.subtotal:,.2f}"
        })
        i += 1
 
    # Conversión a letras
    try:
        total_float = float(orden.total)
        parte_entera = int(total_float)
        parte_decimal = int(round((total_float - parte_entera) * 100))
        letras = num2words(parte_entera, lang='es').upper()
        total_letras = f"{letras} CON {parte_decimal:02d}/100 {nombre_moneda}"
    except Exception:
        total_letras = "---"
 
    # --- C. CONTEXTO PARA LA PLANTILLA HTML ---
    logo_path = os.path.join(app.root_path, 'static', 'img', 'logo.png').replace('\\', '/')
    bancos_path = os.path.join(app.root_path, 'static', 'img', 'bancos.png').replace('\\', '/')
 
    context = {
        'titulo_documento': titulo_doc,
        'codigo_pedido': codigo_visual,
        'mostrar_precios': mostrar_precios,
        'fecha': orden.fecha.strftime("%d/%m/%Y"),
        'cliente_nombre': orden.cliente.nombre,
        'cliente_ruc': orden.cliente.documento,
        'cliente_direccion_fiscal': orden.cliente.direccion,
        'cliente_telefono': orden.cliente.telefono or "",
        'contacto_atte': orden.atencion or "",
        'orden_compra': orden.orden_compra or "",
        'vendedor_celular': celular_texto,
        'tipo_entrega': orden.tipo_entrega,
        'lugar_entrega': orden.direccion_envio,
        'plazo_entrega': texto_entrega,
        'vendedor_nombre': orden.vendedor.nombre_completo,
        'vendedor_cargo': cargo_mostrar,
        'vendedor_email': email_texto,
        'tbl_contents': lista_items,
        'simbolo': simbolo,
        'subtotal_bruto': f"{subtotal_bruto:,.2f}",
        'label_descuento': f"DESCUENTO ({int(orden.descuento_valor)}%)" if orden.descuento_tipo == 'PORCENTAJE' else "DESCUENTO",
        'monto_descuento': f"- {orden.descuento_total:,.2f}",
        'subtotal_neto': f"{orden.subtotal:,.2f}",
        'igv': f"{orden.igv:,.2f}",
        'total': f"{orden.total:,.2f}",
        'son_letras': total_letras,
        'condicion_pago': orden.condicion_pago or "Contado",
        'validez': orden.validez_oferta or "5 días",
        'observacion': orden.observacion or "",
        'logo_path': logo_path,
        'bancos_path': bancos_path,
    }
 
    # --- D. RENDERIZAR HTML CON JINJA2 ---
    html_renderizado = render_template('pdf_cotizacion.html', **context)
 
    # --- E. CONVERTIR HTML -> PDF CON xhtml2pdf ---
    pdf_buffer = io.BytesIO()
    resultado = pisa.CreatePDF(
        src=html_renderizado,
        dest=pdf_buffer,
        encoding='utf-8'
    )
 
    if resultado.err:
        # Si algo falla, te muestro el HTML crudo en el navegador para depurar
        return f"<h2>Error generando PDF (v2)</h2><pre>{html_renderizado}</pre>", 500
 
    pdf_buffer.seek(0)
 
    suffix = "_ALMACEN" if modo == 'almacen' else ""
    nombre_archivo_base = f"{codigo_visual}_{orden.cliente.nombre[:10]}{suffix}_V2"
    nombre_archivo_base = "".join(
        c for c in nombre_archivo_base if c.isalnum() or c in (' ', '.', '-', '_')
    ).strip()
 
    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name=f"{nombre_archivo_base}.pdf",
        mimetype='application/pdf'
    )
 

@app.route('/descargar_nota_pedido/<int:order_id>')
def descargar_nota_pedido(order_id):
    if session.get('role') not in ['admin', 'almacen', 'administracion', 'vendedor']: 
        return "Acceso denegado", 403
        
    orden = Order.query.get_or_404(order_id)
    vendedor = orden.vendedor
    
    titulo_doc = "NOTA DE PEDIDO (ALMACÉN)"
    codigo_visual = f"NP-{orden.id:05d}"
    mostrar_precios = False  # Almacén no necesita ver precios

    texto_entrega = "Inmediata / A coordinar"
    if orden.fecha_entrega:
        texto_entrega = orden.fecha_entrega.strftime("%d/%m/%Y")

    # Procesamiento de Items (Formato HTML limpio para xhtml2pdf)
    lista_items = []
    i = 1
    for d in orden.details:
        sku_final = "SERV"
        if d.product:
            sku_final = d.product.sku
        elif d.item_type == 'FABRICACION':
            mapa_skus = {
                'SERVICIO DE CORTE': 'SRV-CORT', 'SERVICIO DE SOLDADURA': 'SRV-SOLD',
                'SERVICIO DE GALVANIZADO': 'SRV-GALV', 'SERVICIO DE ZINCADO': 'SRV-ZINC',
                'SERVICIO DE ROSCADO': 'SRV-ROSC', 'SERVICIO DE TROPICALIZADO': 'SRV-TROP',
                'SERVICIO GENERAL': 'SRV-GEN'
            }
            titulo_limpio = d.nombre_personalizado_titulo.upper() if d.nombre_personalizado_titulo else ""
            sku_final = mapa_skus.get(titulo_limpio, 'SRV-GEN')
            if sku_final == 'SRV-GEN':
                prod_db = Product.query.filter_by(nombre=titulo_limpio).first()
                if prod_db: sku_final = prod_db.sku
        elif d.item_type == 'GLB':
            sku_final = "GLB-001" 

        # Construir descripción HTML (Reemplazo del RichText viejo)
        if d.product and d.item_type == 'PRODUCTO':
            descripcion_html = d.product.nombre
        else:
            titulo = d.nombre_personalizado_titulo.upper() if d.nombre_personalizado_titulo else ""
            cuerpo = d.nombre_personalizado.upper() if d.nombre_personalizado else ""
            partes = []
            if titulo: partes.append(f"<b>{titulo}</b>")
            if cuerpo: partes.append(cuerpo)
            descripcion_html = " ".join(partes)

        unidad_final = "UND" 
        if d.item_type == 'FABRICACION': unidad_final = "SRV"
        elif d.item_type == 'GLB': unidad_final = "GLB"
        elif d.product and hasattr(d.product, 'unidad_medida'): 
            unidad_final = d.product.unidad_medida or "UND"
        
        ubicacion_final = d.product.ubicacion if (d.product and hasattr(d.product, 'ubicacion')) else ""

        lista_items.append({
            'item': i,
            'sku': sku_final,
            'ubicacion': ubicacion_final,
            'cant': d.cantidad,
            'um': unidad_final,
            'desc': descripcion_html
        })
        i += 1

    # Rutas absolutas para imágenes
    logo_path = os.path.join(app.root_path, 'static', 'img', 'logo.png').replace('\\', '/')

    # Contexto para Jinja2
    context = {
        'titulo_documento': titulo_doc,
        'estado_orden': orden.estado,
        'codigo_pedido': codigo_visual,
        'mostrar_precios': mostrar_precios,
        'fecha': orden.fecha.strftime("%d/%m/%Y"),
        'cliente_nombre': orden.cliente.nombre,
        'cliente_ruc': orden.cliente.documento,
        'cliente_direccion_fiscal': orden.cliente.direccion,
        'cliente_telefono': orden.cliente.telefono or "",
        'contacto_atte': orden.atencion or "",
        'orden_compra': orden.orden_compra or "",
        'tipo_entrega': orden.tipo_entrega,
        'lugar_entrega': orden.direccion_envio,
        'plazo_entrega': texto_entrega,
        'tbl_contents': lista_items,
        'observacion': orden.observacion or "Ninguna",
        'vendedor_nombre': vendedor.nombre_completo,
        'vendedor_usuario': vendedor.username,
        'vendedor_celular': vendedor.celular or "",
        'vendedor_cargo': vendedor.cargo_formal or "Asesor Comercial",
        'vendedor_email': vendedor.email_empresa or "ventas@importbolts.com",
        'almacenero_nombre': orden.almacenero_nombre or "Pendiente",
        'fecha_almacen': orden.fecha_verificacion_almacen.strftime('%d/%m/%Y %H:%M') if orden.fecha_verificacion_almacen else "---",
        'gerente_nombre': orden.gerente_nombre or "Pendiente",
        'fecha_gerencia': orden.fecha_aprobacion.strftime('%d/%m/%Y %H:%M') if orden.fecha_aprobacion else "---",
        'logo_path': logo_path
    }
    
    html_renderizado = render_template('pdf_orden_pedido.html', **context)

    # Generar PDF en memoria
    pdf_buffer = io.BytesIO()
    resultado = pisa.CreatePDF(src=html_renderizado, dest=pdf_buffer, encoding='utf-8')

    if resultado.err:
        return f"<h2>Error generando PDF de Almacén</h2><pre>{html_renderizado}</pre>", 500

    pdf_buffer.seek(0)
    
    nombre_archivo_base = f"{codigo_visual}_{orden.cliente.nombre[:10]}_ALMACEN"
    nombre_archivo_base = "".join([c for c in nombre_archivo_base if c.isalnum() or c in (' ', '.', '-', '_')]).strip()

    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name=f"{nombre_archivo_base}.pdf",
        mimetype='application/pdf'
    )
        

@app.route('/subir_oc/<int:order_id>', methods=['POST'])
def subir_oc(order_id):
    if 'user_id' not in session: 
        return {'status': 'error', 'msg': 'Login requerido'}, 401
    
    orden = Order.query.get_or_404(order_id)
    msg_exito = []

    # NUEVA LÓGICA: Eliminar Archivo
    if request.form.get('eliminar_archivo') == 'SI':
        if orden.archivo_oc:
            # Opcional: Aquí podrías eliminar el archivo físicamente de Amazon S3 usando s3_client.delete_object
            orden.archivo_oc = None
            msg_exito.append("Archivo eliminado")
            db.session.commit()
            return {'status': 'success', 'msg': 'Archivo eliminado correctamente'}
    
    # 1. ACTUALIZAR NÚMERO MANUAL
    if 'numero_oc_manual' in request.form:
        nuevo_numero = request.form.get('numero_oc_manual').strip().upper()
        if nuevo_numero:
            orden.orden_compra = nuevo_numero
            msg_exito.append("Número actualizado")

    # 2. SUBIR ARCHIVO A AMAZON S3
    archivo = request.files.get('archivo_pdf')
    if archivo and archivo.filename != '':
        # Validar extensión
        if not archivo.filename.lower().endswith(('.pdf', '.jpg', '.jpeg', '.png')):
            return {'status': 'error', 'msg': 'Formato no válido (Use PDF o Imagen)'}

        # Nombre limpio
        ext = archivo.filename.split('.')[-1]
        nombre_limpio = secure_filename(f"OC_{orden.id:05d}_{orden.cliente.nombre[:10]}.{ext}")
        
        try:
            # Subir a S3 en lugar de disco local
            s3_client.upload_fileobj(
                archivo,
                S3_BUCKET_NAME,
                nombre_limpio,
                ExtraArgs={"ContentType": archivo.content_type} # Permite previsualizar en el navegador
            )
            orden.archivo_oc = nombre_limpio
            msg_exito.append("Archivo subido")
            
        except Exception as e:
            return {'status': 'error', 'msg': f'Fallo al subir a la nube: {str(e)}'}

    try:
        db.session.commit()
        return {'status': 'success', 'msg': " y ".join(msg_exito) if msg_exito else "Sin cambios"}
    except Exception as e:
        db.session.rollback()
        return {'status': 'error', 'msg': f'Error de Base de datos: {str(e)}'}


@app.route('/ver_oc/<filename>')
def ver_oc(filename):
    try:
        # 1. El servidor se conecta a Amazon y lee el archivo en memoria
        # Esto usa el mismo método exacto que la subida (que ya sabemos que funciona perfecto)
        archivo_s3 = s3_client.get_object(Bucket=S3_BUCKET_NAME, Key=filename)

        # 2. Le enviamos el archivo directamente a la pantalla del usuario
        return send_file(
            io.BytesIO(archivo_s3['Body'].read()),
            mimetype='application/pdf',
            as_attachment=False,  # False = Se abre como pestaña nueva. True = Fuerza descarga directa
            download_name=filename
        )
    except Exception as e:
        return f"<h1>Error al recuperar el documento desde la nube</h1><p>{str(e)}</p>", 404

@app.route('/categoria/nueva', methods=['POST'])
def nueva_categoria():
    if session.get('role') not in ['admin', 'almacen']: return "No autorizado", 403
    
    nombre = request.form.get('cat_nombre', '').strip().upper()
    prefijo = request.form.get('cat_prefijo', '').strip().upper()
    
    if not nombre or not prefijo:
        flash('Error: Nombre y Prefijo son obligatorios')
        return redirect(url_for('inventario'))
        
    # Validar duplicados
    if Category.query.filter_by(nombre=nombre).first():
        flash('Error: Esa familia ya existe.')
        return redirect(url_for('inventario'))
        
    if Category.query.filter_by(prefijo=prefijo).first():
        flash(f'Error: El prefijo {prefijo} ya está en uso.')
        return redirect(url_for('inventario'))
    
    try:
        nueva = Category(nombre=nombre, prefijo=prefijo, contador=0)
        db.session.add(nueva)
        db.session.commit()
        flash(f'✅ Familia "{nombre}" creada. Ahora puede seleccionarla en Nuevo Producto.')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}')
        
    # Volvemos al inventario
    return redirect(url_for('inventario'))

# --- ACTUALIZAR ESTA FUNCIÓN EN APP.PY ---
@app.route('/categoria/eliminar', methods=['POST'])
def eliminar_categoria():
    if session.get('role') != 'admin': 
        return {'status': 'error', 'msg': 'No autorizado'}, 403
    
    cat_nombre = request.form.get('nombre_cat')
    
    # 1. Seguridad: Verificar productos asociados
    productos_asociados = Product.query.filter_by(categoria=cat_nombre).count()
    
    if productos_asociados > 0:
        return {
            'status': 'error', 
            'msg': f'⛔ No se puede eliminar "{cat_nombre}". Tiene {productos_asociados} productos asociados.'
        }
    
    # 2. Borrar si está vacía
    cat_a_borrar = Category.query.filter_by(nombre=cat_nombre).first()
    if cat_a_borrar:
        try:
            db.session.delete(cat_a_borrar)
            db.session.commit()
            return {'status': 'success', 'msg': f'Familia "{cat_nombre}" eliminada.'}
        except Exception as e:
            db.session.rollback()
            return {'status': 'error', 'msg': str(e)}
    
    return {'status': 'error', 'msg': 'La familia no existe.'}



@app.route('/api/listar_todos_clientes')
def listar_todos_clientes():
    if session.get('user_id') is None: return {'data': []}
    
    # Traemos todos los clientes ordenados por la última vez que se actualizaron
    clientes = Client.query.order_by(Client.last_updated.desc()).all()
    
    data = []
    for c in clientes:
        data.append({
            'documento': c.documento,
            'nombre': c.nombre,
            'direccion': c.direccion or '',
            'telefono': c.telefono or '',
            'estado': c.estado,
            'condicion': c.condicion,
            'updated_at': c.last_updated.strftime('%d/%m/%Y %H:%M'),
            'updated_by': c.updated_by
        })
    
    return {'data': data}

# --- 1. API PARA PREVISUALIZAR LA LISTA COMPLETA ---
@app.route('/api/preview_minimos', methods=['POST'])
def preview_minimos():
    if session.get('role') != 'admin': return {'status': 'error', 'msg': 'No autorizado'}, 403
    
    familia = request.form.get('categoria_nombre')
    calidad = request.form.get('calidad_nombre') # Opcional
    
    # Construir la consulta base
    query = Product.query.filter_by(categoria=familia)
    
    # Si seleccionó calidad, filtramos más
    if calidad and calidad != 'TODAS':
        query = query.filter_by(calidad=calidad)
        
    productos = query.order_by(Product.sku.asc()).all()
    
    # Armar lista completa para la tabla
    lista = []
    for p in productos:
        lista.append({
            'id': p.id,
            'sku': p.sku,
            'nombre': p.nombre,
            'min_actual': p.stock_minimo
        })
    
    return {
        'status': 'success',
        'total': len(lista),
        'productos': lista
    }

# --- 2. RUTA DE ACTUALIZACIÓN MASIVA (POR IDs SELECCIONADOS) ---
@app.route('/config/minimos_masivos', methods=['POST'])
def actualizar_minimos_masivos():
    # Permitimos a admin y almacen (que son los que tienen el botón)
    if session.get('role') not in ['admin', 'almacen']: 
        return {'status': 'error', 'msg': 'No autorizado'}, 403
    
    data = request.get_json()
    if not data:
        return {'status': 'error', 'msg': 'No se recibieron datos correctamente.'}

    ids = data.get('ids', [])
    
    try:
        nuevo_minimo = int(data.get('nuevo_minimo'))
        if nuevo_minimo < 0: raise ValueError("No negativos")
    except:
        return {'status': 'error', 'msg': 'Cantidad de stock mínimo inválida.'}

    if not ids:
        return {'status': 'error', 'msg': 'Debe seleccionar al menos un producto.'}

    try:
        # Ejecutar Update Masivo SOLO a los IDs seleccionados
        resultado = Product.query.filter(Product.id.in_(ids)).update(
            {Product.stock_minimo: nuevo_minimo}, 
            synchronize_session=False
        )
        
        db.session.commit()
        return {'status': 'success', 'msg': f'Se actualizaron {resultado} productos al nuevo mínimo de {nuevo_minimo}.'}
        
    except Exception as e:
        db.session.rollback()
        return {'status': 'error', 'msg': str(e)}

@app.route('/producto/importar', methods=['POST'])
def importar_excel():
    import gc
    import traceback
    from openpyxl import load_workbook

    if session.get('role') not in ['admin', 'almacen']:
        return "No autorizado", 403

    if 'archivo_excel' not in request.files:
        flash('No se seleccionó ningún archivo')
        return redirect(url_for('inventario'))

    archivo = request.files['archivo_excel']
    if not archivo or archivo.filename == '':
        return redirect(url_for('inventario'))

    filename = secure_filename(archivo.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    archivo.save(filepath)

    nuevos = 0
    actualizados = 0
    errores = 0

    try:
        # ================================================================
        # OPENPYXL EN MODO STREAMING (read_only=True)
        # Lee una fila a la vez sin cargar el archivo completo en RAM
        # Un Excel de 7MB con 4000 filas usa solo ~5MB en vez de 500MB
        # ================================================================
        wb = load_workbook(filename=filepath, read_only=True, data_only=True)

        # Intentar hoja 'STOCK', si no existe usar la primera
        if 'STOCK' in wb.sheetnames:
            ws = wb['STOCK']
        else:
            ws = wb.active

        # Leer encabezados de la primera fila
        headers = []
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip().upper() if h is not None else '' for h in header_row]

        print(f">>> [IMPORTAR] Columnas detectadas: {headers}")

        # Mapeo flexible de columnas
        def get_col(row_vals, *posibles_nombres):
            for nombre in posibles_nombres:
                if nombre in headers:
                    idx = headers.index(nombre)
                    if idx < len(row_vals):
                        return row_vals[idx]
            return None

        hora_actual = hora_peru()
        usuario_actual = session.get('username', 'Sistema')
        user_id_actual = session.get('user_id')

        # ================================================================
        # PRE-CARGAR CACHÉ (evita SELECT por cada fila)
        # ================================================================
        skus_existentes = {p.sku: p.id for p in db.session.query(Product.id, Product.sku).all()}
        cats_existentes = {c.nombre: c.prefijo for c in db.session.query(Category.nombre, Category.prefijo).all()}
        db.session.expunge_all()
        gc.collect()

        # ================================================================
        # PROCESAR FILA POR FILA (modo streaming, sin acumular en RAM)
        # ================================================================
        batch_updates = []
        batch_inserts = []
        batch_kardex  = []
        BATCH_SIZE = 100

        fila_num = 0
        for row_vals in ws.iter_rows(min_row=2, values_only=True):
            fila_num += 1

            # Leer SKU
            sku_raw = get_col(row_vals, 'CÓDIGO', 'CODIGO', 'SKU', 'CÓDIGO ')
            if sku_raw is None:
                continue
            sku = str(sku_raw).strip()
            if sku.endswith('.0'):
                sku = sku[:-2]
            if not sku or sku.lower() in ('nan', 'none', ''):
                continue

            # Leer resto de campos
            def clean_str(val, default=''):
                if val is None: return default
                s = str(val).strip()
                return default if s.lower() in ('nan', 'none', '') else s

            def clean_int(val, default=0):
                try:
                    v = str(val).replace(',', '').strip()
                    return int(float(v)) if v and v.lower() not in ('nan', 'none', '') else default
                except:
                    return default

            nombre   = clean_str(get_col(row_vals, 'DESCRIPCIÓN', 'DESCRIPCION', 'NOMBRE'), 'Sin Nombre')
            familia  = clean_str(get_col(row_vals, 'FAMILIA', 'CATEGORIA'), 'GENERAL')
            calidad  = clean_str(get_col(row_vals, 'CALIDAD'), '-')
            ubicacion= clean_str(get_col(row_vals, 'UBICACION', 'UBICACIÓN'))
            estado_v = clean_str(get_col(row_vals, 'ESTADO')).upper()
            if estado_v == 'OK': estado_v = ''

                        # --- TRUNCADO DE SEGURIDAD (evita StringDataRightTruncation) ---
            # Limpia caracteres raros/corruptos y limita longitud
            import unicodedata

            def limpiar_campo(valor, max_len):
                if not valor:
                    return ''
                # Quitar caracteres de control y símbolos raros
                limpio = ''.join(
                    c for c in str(valor)
                    if unicodedata.category(c)[0] not in ('C',)  # Elimina caracteres de control
                    and ord(c) < 65536  # Elimina emojis/símbolos raros
                ).strip()
                return limpio[:max_len]

            nombre    = limpiar_campo(nombre, 490)
            familia   = limpiar_campo(familia, 190)
            calidad   = limpiar_campo(calidad, 190)
            ubicacion = limpiar_campo(ubicacion, 190)
            estado_v  = limpiar_campo(estado_v, 90)

            stock_val = clean_int(get_col(row_vals, 'CANT. ACT.', 'STOCK', 'CANTIDAD', 'CANT.ACT.'))
            min_val   = clean_int(get_col(row_vals, 'STOCK MÍNIMO', 'STOCK MINIMO', 'MINIMO'), 10)

            # --- LEER PRECIO UNITARIO (Si existe, si no queda en 0.0) ---
            def clean_float(val, default=0.0):
                try:
                    v = str(val).replace(',', '').strip()
                    return float(v) if v and v.lower() not in ('nan', 'none', '') else default
                except:
                    return default

            precio_unit = clean_float(get_col(row_vals, 'PRECIO UNI.','PRECIO UNIT', 'PRECIO UNIDAD', 'P. UNIT', 'PRECIO_UNIT', 'PRECIO UNITARIO'))

            # Crear categoría si no existe
            if familia not in cats_existentes:
                base = "".join(c for c in familia[:3].upper() if c.isalnum()) or "GEN"
                prefijo_final = base
                n = 1
                prefijos_usados = set(cats_existentes.values())
                while prefijo_final in prefijos_usados:
                    prefijo_final = f"{base[:2]}{n}"
                    n += 1
                nuevo_cat = Category(nombre=familia, prefijo=prefijo_final, contador=0)
                db.session.add(nuevo_cat)
                db.session.flush()
                cats_existentes[familia] = prefijo_final

            # Decidir INSERT o UPDATE
            if sku in skus_existentes:
                upd = {
                    'sku': sku, 'nombre': nombre, 'categoria': familia,
                    'calidad': calidad, 'ubicacion': ubicacion, 'estado': estado_v,
                    'stock_actual': stock_val, 'stock_minimo': min_val,
                    'fecha_actualizacion': hora_actual, 'actualizado_por': usuario_actual
                }
                # Solo actualiza el precio si el Excel trae un valor mayor a 0
                if precio_unit > 0:
                    upd['tiene_precio'] = True
                    upd['precio_unidad'] = precio_unit
                else:
                    upd['tiene_precio'] = False
                    upd['precio_unidad'] = 0.0

                batch_updates.append(upd)
                actualizados += 1
            else:
                nuevo_prod = Product(
                    sku=sku, nombre=nombre, categoria=familia, calidad=calidad,
                    ubicacion=ubicacion, stock_actual=stock_val, stock_minimo=min_val,
                    precio_unidad=precio_unit,          # <-- Ahora toma el valor del Excel
                    precio_caja=0.0,
                    precio_docena=precio_unit,          # <-- Docena igual al unitario por defecto
                    costo_referencial=0.0, estado=estado_v,
                    fecha_actualizacion=hora_actual, actualizado_por=usuario_actual
                )
                batch_inserts.append(nuevo_prod)
                skus_existentes[sku] = -1
                nuevos += 1

            # Cada BATCH_SIZE filas: commit y limpiar
            if (nuevos + actualizados) % BATCH_SIZE == 0:
                # Ejecutar updates
                for upd in batch_updates:
                    if upd.get('tiene_precio'):
                        db.session.execute(text("""
                            UPDATE product SET
                                nombre=:nombre, categoria=:categoria, calidad=:calidad,
                                ubicacion=:ubicacion, estado=:estado,
                                stock_actual=:stock_actual, stock_minimo=:stock_minimo,
                                precio_unidad=:precio_unidad,
                                precio_docena=:precio_unidad,
                                fecha_actualizacion=:fecha_actualizacion,
                                actualizado_por=:actualizado_por
                            WHERE sku=:sku
                        """), upd)
                    else:
                        db.session.execute(text("""
                            UPDATE product SET
                                nombre=:nombre, categoria=:categoria, calidad=:calidad,
                                ubicacion=:ubicacion, estado=:estado,
                                stock_actual=:stock_actual, stock_minimo=:stock_minimo,
                                fecha_actualizacion=:fecha_actualizacion,
                                actualizado_por=:actualizado_por
                            WHERE sku=:sku
                        """), upd)

                # Insertar nuevos
                if batch_inserts:
                    db.session.add_all(batch_inserts)
                    db.session.flush()
                    for p in batch_inserts:
                        if p.stock_actual > 0 and p.id:
                            batch_kardex.append(ProductMovement(
                                product_id=p.id, user_id=user_id_actual,
                                tipo='ENTRADA', cantidad=p.stock_actual,
                                stock_anterior=0, stock_nuevo=p.stock_actual,
                                motivo="Saldo Inicial (Importación)"
                            ))
                    if batch_kardex:
                        db.session.add_all(batch_kardex)

                db.session.commit()
                db.session.expunge_all()

                batch_updates = []
                batch_inserts = []
                batch_kardex  = []
                gc.collect()
                print(f">>> [BATCH] Commit: {nuevos + actualizados} filas procesadas")

        # Procesar el último batch (filas restantes)
        for upd in batch_updates:
            if upd.get('tiene_precio'):
                db.session.execute(text("""
                    UPDATE product SET
                        nombre=:nombre, categoria=:categoria, calidad=:calidad,
                        ubicacion=:ubicacion, estado=:estado,
                        stock_actual=:stock_actual, stock_minimo=:stock_minimo,
                        precio_unidad=:precio_unidad,
                        precio_docena=:precio_unidad,
                        fecha_actualizacion=:fecha_actualizacion,
                        actualizado_por=:actualizado_por
                    WHERE sku=:sku
                """), upd)
            else:
                db.session.execute(text("""
                    UPDATE product SET
                        nombre=:nombre, categoria=:categoria, calidad=:calidad,
                        ubicacion=:ubicacion, estado=:estado,
                        stock_actual=:stock_actual, stock_minimo=:stock_minimo,
                        fecha_actualizacion=:fecha_actualizacion,
                        actualizado_por=:actualizado_por
                    WHERE sku=:sku
                """), upd)

        if batch_inserts:
            db.session.add_all(batch_inserts)
            db.session.flush()
            for p in batch_inserts:
                if p.stock_actual > 0 and p.id:
                    batch_kardex.append(ProductMovement(
                        product_id=p.id, user_id=user_id_actual,
                        tipo='ENTRADA', cantidad=p.stock_actual,
                        stock_anterior=0, stock_nuevo=p.stock_actual,
                        motivo="Saldo Inicial (Importación)"
                    ))
            if batch_kardex:
                db.session.add_all(batch_kardex)

        db.session.commit()
        db.session.expunge_all()

        # Cerrar workbook y liberar
        wb.close()
        del wb, skus_existentes, cats_existentes
        del batch_updates, batch_inserts, batch_kardex
        gc.collect()

        # Registrar última importación
        config_import = SystemConfig.query.get('ultima_importacion')
        hora_final = hora_peru()
        if not config_import:
            config_import = SystemConfig(
                key='ultima_importacion', value='EXITOSO',
                updated_at=hora_final, updated_by=usuario_actual
            )
            db.session.add(config_import)
        else:
            config_import.updated_at = hora_final
            config_import.updated_by = usuario_actual
        db.session.commit()

        flash(f'✅ Importación completada: {nuevos} nuevos, {actualizados} actualizados.')

    except Exception as e:
        db.session.rollback()
        print(f"ERROR IMPORTACIÓN:\n{traceback.format_exc()}")
        flash(f'Error en la importación: {str(e)}')
    finally:
        if os.path.exists(filepath):
            os.remove(filepath)
        gc.collect()

    return redirect(url_for('inventario'))


@app.route('/fix_columnas_secreto_2026')
def fix_columnas():
    try:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE product ALTER COLUMN nombre TYPE VARCHAR(500)"))
            conn.execute(text("ALTER TABLE product ALTER COLUMN categoria TYPE VARCHAR(200)"))
            conn.execute(text("ALTER TABLE product ALTER COLUMN calidad TYPE VARCHAR(200)"))
            conn.execute(text("ALTER TABLE product ALTER COLUMN ubicacion TYPE VARCHAR(200)"))
            conn.execute(text("ALTER TABLE product ALTER COLUMN estado TYPE VARCHAR(200)"))
            conn.commit()
        return "<h2>✅ Columnas ampliadas correctamente. Ya puedes importar el Excel grande.</h2>"
    except Exception as e:
        return f"<h2>Error: {str(e)}</h2>"

# 2. ACTUALIZAR NUEVO PRODUCTO (Para responder JSON y no borrar datos)
# --- FUNCIÓN NUEVO PRODUCTO (Actualizada) ---
@app.route('/producto/nuevo', methods=['POST'])
def nuevo_producto():
    if session.get('role') not in ['admin', 'almacen']: 
        return {'status': 'error', 'msg': 'No autorizado'}, 403
    
    try:
        sku_manual = request.form.get('sku', '').strip()
        familia_nombre = request.form['categoria_nombre']
        nombre = request.form['nombre'].strip()
        calidad = request.form['calidad'].strip()
        ubicacion = request.form.get('ubicacion', '').strip()
        estado_val = request.form.get('estado', '').strip().upper()
        if estado_val == 'OK': estado_val = ""
        
        try:
            stock = int(request.form['stock'])
            # NUEVO: Capturar Stock Mínimo
            stock_min = int(request.form.get('stock_minimo', 10))
            p_unidad = float(request.form['p_unidad']) if request.form['p_unidad'] else 0.0
            p_caja = float(request.form['p_caja']) if request.form['p_caja'] else 0.0
        except:
            return {'status': 'error', 'msg': 'Formato numérico inválido'}

        if not nombre: return {'status': 'error', 'msg': 'Falta la descripción'}
        if stock < 0 or p_unidad < 0 or stock_min < 0: return {'status': 'error', 'msg': 'No negativos'}

        sku_final = ""
        if sku_manual:
            sku_final = sku_manual.upper()
            if Product.query.filter_by(sku=sku_final).first():
                return {'status': 'error', 'msg': f'El SKU "{sku_final}" ya existe.'}
        else:
            cat = Category.query.filter_by(nombre=familia_nombre).first()
            if not cat:
                base = "".join(c for c in familia_nombre[:3].upper() if c.isalnum()) or "GEN"
                cat = Category(nombre=familia_nombre, prefijo=base, contador=0)
                db.session.add(cat)
            cat.contador += 1
            sku_final = f"{cat.prefijo}-{str(cat.contador).zfill(4)}"

        nuevo = Product(
            sku=sku_final, nombre=nombre, categoria=familia_nombre, calidad=calidad,
            ubicacion=ubicacion, stock_actual=stock, stock_minimo=stock_min, # Guardar
            precio_unidad=p_unidad, precio_caja=p_caja, precio_docena=p_unidad * 0.9, costo_referencial=0.0
        )
        db.session.add(nuevo)
        db.session.flush()
        
        if stock > 0:
            kardex = ProductMovement(
                product_id=nuevo.id, user_id=session['user_id'], tipo='ENTRADA',
                cantidad=stock, stock_anterior=0, stock_nuevo=stock, motivo="Saldo Inicial"
            )
            db.session.add(kardex)

        registrar_log(f"Creó producto {sku_final}", "bi-plus-circle-fill", "text-success")
        db.session.commit()
        
        return {'status': 'success', 'msg': 'Creado', 'sku': sku_final}
        
    except Exception as e:
        db.session.rollback()
        return {'status': 'error', 'msg': str(e)}

# --- FUNCIÓN EDITAR PRODUCTO (Actualizada) ---
@app.route('/producto/editar', methods=['POST'])
def editar_producto():
    if session.get('role') != 'admin': return "Acceso denegado", 403
    
    try:
        prod_id = request.form['prod_id']
        url_origen = request.form.get('url_origen')
        
        prod = Product.query.get(prod_id)
        if not prod:
            flash('Producto no encontrado')
            return redirect(url_for('inventario'))

        nombre = request.form['nombre'].strip()
        nueva_familia = request.form.get('categoria', '').strip()
        nueva_calidad = request.form.get('calidad', '').strip()
        estado_val = request.form.get('estado', '').strip().upper()
        if estado_val == 'OK': estado_val = ""
        
        if not nombre:
            flash('⛔ Error: La descripción no puede estar vacía.')
            return redirect(url_for('inventario'))
        if not nueva_familia or not nueva_calidad:
            flash('⛔ Error: Familia y Calidad son obligatorias.')
            return redirect(url_for('inventario'))

        prod.nombre = nombre
        # El stock actual NO se edita aquí, solo el mínimo
        # NUEVO: Actualizar Stock Mínimo
        prod.stock_minimo = int(request.form.get('stock_minimo', 10))
        prod.precio_unidad = float(request.form['p_unidad'])
        prod.precio_caja = float(request.form['p_caja'])
        prod.ubicacion = request.form.get('ubicacion', '').strip()
        prod.categoria = nueva_familia
        prod.calidad = nueva_calidad
        prod.estado = estado_val
        
        registrar_log(f"Editó producto {prod.sku}", "bi-pencil-fill", "text-warning")
        
        db.session.commit()
        flash('✅ Producto actualizado correctamente.')
        
        if url_origen: return redirect(url_origen)
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error al editar: {str(e)}')
        
    return redirect(url_for('inventario'))

# --- FUNCIÓN ELIMINAR INTELIGENTE ---
@app.route('/producto/eliminar/<int:prod_id>')
def eliminar_producto(prod_id):
    if session.get('role') != 'admin': 
        flash('No tiene permisos para eliminar.')
        return redirect(url_for('inventario'))
    
    try:
        # 1. Obtener el producto
        prod = Product.query.get_or_404(prod_id)
        sku_eliminado = prod.sku

        # 2. VERIFICACIÓN DE SEGURIDAD (Ventas)
        # Si el producto ya se vendió, PROHIBIDO eliminarlo.
        # (Asumiendo que tienes una relación 'ventas_detalle' o consultando OrderDetail)
        # Si no has definido la relación en models, importamos y consultamos:
        from models import OrderDetail, ProductMovement # Asegúrate de importar esto arriba
        
        ventas = OrderDetail.query.filter_by(product_id=prod_id).first()
        if ventas:
            flash(f'⛔ No se puede eliminar {sku_eliminado}: Ya tiene ventas registradas. Use "Desactivar" o ajuste el stock a 0.')
            # Redirigir a la página anterior (Mantiene filtros)
            return redirect(request.referrer or url_for('inventario'))

        # 3. LIMPIEZA DE KARDEX (Solo si no hay ventas)
        # Borramos sus movimientos de stock (Ingreso inicial, ajustes, etc.)
        ProductMovement.query.filter_by(product_id=prod_id).delete()

        # 4. ELIMINAR EL PRODUCTO FINALMENTE
        db.session.delete(prod)
        db.session.commit()
        
        flash(f'✅ Producto {sku_eliminado} y su historial inicial han sido eliminados.')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar: {str(e)}')

    # 5. RETORNO INTELIGENTE
    # request.referrer te devuelve a la URL exacta donde estabas (página 2, búsqueda "SEN", etc.)
    return redirect(request.referrer or url_for('inventario'))

@app.route('/producto/ajustar_stock', methods=['POST'])
def ajustar_stock():
    if session.get('role') not in ['admin', 'almacen']: return "No autorizado", 403
    
    prod_id = request.form['prod_id']
    tipo_ajuste = request.form['tipo']
    cantidad = int(request.form['cantidad'])
    motivo_texto = request.form['motivo']
    
    # CAPTURAR LA URL DE RETORNO (Aquí está la magia)
    url_origen = request.form.get('url_origen')

    prod = Product.query.get(prod_id)
    stock_antes = prod.stock_actual
    
    tipo_kardex = ""
    
    if tipo_ajuste == 'ingreso':
        prod.stock_actual += cantidad
        tipo_kardex = "ENTRADA"
        flash(f'Ingreso registrado: +{cantidad} en {prod.sku}')
    else:
        prod.stock_actual -= cantidad
        tipo_kardex = "SALIDA"
        flash(f'Salida registrada: -{cantidad} en {prod.sku}')
        
    kardex = ProductMovement(
        product_id=prod.id,
        user_id=session['user_id'],
        tipo=tipo_kardex,
        cantidad=cantidad,
        stock_anterior=stock_antes,
        stock_nuevo=prod.stock_actual,
        motivo=motivo_texto
    )
    db.session.add(kardex)
    db.session.commit()
    
    # SI TENEMOS URL DE ORIGEN, VOLVEMOS AHÍ. SI NO, AL INICIO.
    if url_origen:
        return redirect(url_origen)
    
    return redirect(url_for('inventario'))

@app.route('/kardex')
def ver_kardex():
    if session.get('user_id') is None: return redirect(url_for('login'))
    
    # Unimos con Product para poder filtrar por nombre/categoría
    query = ProductMovement.query.join(Product)
    
    # 1. Filtro por Texto (Nombre, SKU o Motivo)
    busqueda = request.args.get('busqueda')
    if busqueda:
        query = query.filter(
            or_(
                Product.nombre.ilike(f"%{busqueda}%"),
                Product.sku.ilike(f"%{busqueda}%"),
                ProductMovement.motivo.ilike(f"%{busqueda}%")
            )
        )
    
    # 2. Filtro por Categoría
    cat_nombre = request.args.get('categoria')
    if cat_nombre and cat_nombre != 'todas':
        query = query.filter(Product.categoria == cat_nombre)

    # 3. Filtro por Tipo (Entrada/Salida)
    tipo_mov = request.args.get('tipo')
    if tipo_mov and tipo_mov in ['ENTRADA', 'SALIDA']:
        query = query.filter(ProductMovement.tipo == tipo_mov)

    # ---> NUEVO: FILTRO PARA OCULTAR SALDOS INICIALES <---
    ocultar_iniciales = request.args.get('ocultar_iniciales')
    if ocultar_iniciales == 'on':
        # Filtramos excluyendo (~) los motivos que contengan la palabra "Inicial"
        query = query.filter(~ProductMovement.motivo.ilike('%Inicial%'))

    # 4. Filtro por Rango de Fechas
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    
    if fecha_inicio and fecha_fin:
        start = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        end = datetime.strptime(fecha_fin + " 23:59:59", '%Y-%m-%d %H:%M:%S')
        query = query.filter(ProductMovement.fecha.between(start, end))
        
    # --- NUEVO: PAGINACIÓN (En vez del limit) ---
    query = query.order_by(ProductMovement.fecha.desc())
    
    page = request.args.get('page', 1, type=int)
    per_page = 25 # Muestra 25 movimientos por página
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    
    movimientos = pagination.items
    categorias = Category.query.all()
    
    return render_template('kardex.html', 
                           movimientos=movimientos, 
                           categorias=categorias,
                           pagination=pagination) # Pasamos la info de las páginas al HTML

@app.route('/despachos')
def despachos():
    # Permitimos acceso a admin y almacen (y chofer si quieres que vean su historial)
    if session.get('role') not in ['admin', 'almacen', 'chofer']: return "Acceso denegado", 403
    
    vista = request.args.get('vista', 'pendientes')
    query = Order.query
    
    # 1. Filtros básicos por estado
    if vista == 'pendientes':
        query = query.filter(Order.estado == 'Aprobado')
    elif vista == 'proceso':
        query = query.filter(Order.estado == 'En Preparacion')
    elif vista == 'finalizados':
        query = query.filter(Order.estado.in_(['Despachado', 'Entregado']))

    # 2. Ordenamiento
    modo_orden = request.args.get('ordenar_por', 'urgencia')
    if modo_orden == 'urgencia':
        query = query.order_by(Order.fecha_entrega.asc())
    else:
        query = query.order_by(Order.fecha.asc())

    # 3. Límite para historial
    if vista == 'finalizados':
        query = query.limit(50)

    ordenes = query.all()
    
    # --- 🔴 NUEVO: CARGAR LA LISTA DE CHOFERES ---
    # Esto es lo que te faltaba. Busca todos los usuarios con rol 'chofer'
    choferes = User.query.filter_by(role='chofer').all()
    
    # Contadores
    count_pend = Order.query.filter(Order.estado == 'Aprobado').count()
    count_proc = Order.query.filter(Order.estado == 'En Preparacion').count()
    
    return render_template('despachos.html', 
                           ordenes=ordenes, 
                           vista_actual=vista,
                           orden_actual=modo_orden,
                           c_pend=count_pend,
                           c_proc=count_proc,
                           hoy=hora_peru().date(),
                           choferes=choferes) # <--- 🔴 IMPORTANTE: ENVIARLOS AQUI

# --- EN APP.PY ---

@app.route('/logistica/cambiar_tipo_entrega/<int:order_id>')
def cambiar_tipo_entrega(order_id):
    if session.get('role') not in ['admin', 'almacen']: return "Acceso denegado", 403
    
    orden = Order.query.get_or_404(order_id)
    
    # Lógica de switch simple
    if orden.tipo_entrega == 'Envio':
        orden.tipo_entrega = 'Recojo'
        # Limpiamos dirección si quieres, o la dejas por si se arrepiente de nuevo
        flash(f'Orden #{order_id} cambiada a RECOJO EN TIENDA.')
    else:
        orden.tipo_entrega = 'Envio'
        flash(f'Orden #{order_id} cambiada a ENVÍO A DOMICILIO.')
    
    db.session.commit()
    
    # Volvemos a la misma pantalla de despachos
    return redirect(url_for('despachos', vista='proceso'))

# --- NUEVA RUTA: CAMBIAR A "EN PREPARACIÓN" ---
@app.route('/iniciar_picking/<int:order_id>')
def iniciar_picking(order_id):
    if session.get('role') not in ['admin', 'almacen']: return "Acceso denegado", 403
    
    orden = Order.query.get_or_404(order_id)
    if orden.estado == 'Aprobado':
        orden.estado = 'En Preparacion'
        # Aquí podrías guardar quién lo inició: orden.almacenero_id = session['user_id']
        db.session.commit()
        
    return redirect(url_for('despachos', vista='proceso'))

@app.route('/cobranzas')
def cobranzas():
    # AHORA: Solo Admin y Administración (Vendedores NO, Almacén NO)
    if session.get('role') not in ['admin', 'administracion']: return "Acceso denegado", 403
    
    filtro = request.args.get('ver', 'deudas')
    query = Order.query
    if filtro == 'deudas':
        query = query.filter(Order.estado_pago != 'Pagado')
    ordenes = query.order_by(Order.fecha.asc()).all()
    return render_template('cobranzas.html', ordenes=ordenes)

@app.route('/registrar_pago', methods=['POST'])
def registrar_pago():
    if session.get('role') not in ['admin', 'administracion']: return "Acceso denegado", 403
    
    order_id = request.form['order_id']
    monto = float(request.form['monto'])
    metodo = request.form['metodo']
    nota = request.form['nota']
    
    orden = Order.query.get(order_id)
    
    # Validar que no pague más de la deuda
    deuda_actual = orden.total - orden.monto_pagado
    if monto > (deuda_actual + 0.1): # Margen de error 0.1 por decimales
        flash('Error: El monto excede la deuda actual.')
        return redirect(url_for('cobranzas'))
    
    # 1. Crear registro de pago
    nuevo_pago = Payment(
        order_id=orden.id,
        monto=monto,
        metodo=metodo,
        nota=nota,
        fecha=hora_peru()
    )
    db.session.add(nuevo_pago)
    
    # 2. Actualizar la Orden
    orden.monto_pagado += monto
    
    # Calcular nuevo estado
    if orden.monto_pagado >= (orden.total - 0.1):
        orden.estado_pago = 'Pagado'
        orden.monto_pagado = orden.total # Ajuste exacto
    elif orden.monto_pagado > 0:
        orden.estado_pago = 'Parcial'
    else:
        orden.estado_pago = 'Pendiente'
        
    db.session.commit()
    flash(f'Pago de S/ {monto} registrado correctamente.')
    return redirect(url_for('cobranzas'))

# --- REEMPLAZAR ESTA FUNCIÓN EN APP.PY ---

@app.route('/cambiar_estado/<int:order_id>/<nuevo_estado>')
def cambiar_estado(order_id, nuevo_estado):
    if 'user_id' not in session: return {'status': 'error', 'msg': 'Login requerido'}, 401
    
    orden = Order.query.get_or_404(order_id)
    orden.estado = nuevo_estado
    
    # --- NUEVO: GUARDAR DATOS DE ALMACÉN (Si se enviaron) ---
    peso = request.args.get('peso_kardex')
    bultos = request.args.get('bultos')
    
    if peso: orden.peso_total = peso # Asegúrate de tener esta columna en tu modelo
    if bultos: orden.cantidad_bultos = bultos # Asegúrate de tener esta columna en tu modelo
    
    db.session.commit()
    
    # --- NUEVO: SISTEMA DE REDIRECCIÓN ---
    # Si la petición viene con ?origin=despachos, hacemos redirect en vez de JSON
    origen = request.args.get('origin')
    
    if origen == 'despachos':
        # Si despachamos o entregamos, vamos a la pestaña de finalizados
        if nuevo_estado in ['Despachado', 'Entregado']:
            return redirect(url_for('despachos', vista='finalizados'))
        else:
            return redirect(url_for('despachos', vista='proceso'))
            
    # Si viene de Historial Ventas (AJAX), devolvemos JSON como siempre
    return {'status': 'success', 'msg': f'Estado actualizado a {nuevo_estado}'}

@app.route('/reportes_predicciones')
def reportes_predicciones():
    if session.get('role') not in ['admin', 'administracion']: return "Acceso denegado", 403
    
    # 1. CALCULO DE PREDICCIONES POR PRODUCTO
    # Obtenemos productos que han tenido movimiento de SALIDA (Ventas)
    productos_activos = db.session.query(
        Product.nombre, 
        Product.stock_actual,
        func.sum(ProductMovement.cantidad).label('total_vendido')
    ).join(ProductMovement).filter(
        ProductMovement.tipo == 'SALIDA',
        # Analizamos los últimos 90 días (Trimestre) para mejor precisión
        ProductMovement.fecha >= hora_peru() - timedelta(days=90)
    ).group_by(Product.id).all()
    
    reporte = []
    
    for p in productos_activos:
        # Promedio mensual real (basado en los ultimos 3 meses)
        promedio_mensual = p.total_vendido / 3 
        
        # Algoritmo Simple de Predicción:
        # Asumimos un crecimiento del 10% o estacionalidad
        prediccion = promedio_mensual * 1.10
        
        estado_proyeccion = "Estable"
        if prediccion > p.stock_actual:
            estado_proyeccion = "QUIEBRE DE STOCK (Comprar urgente)"
        
        reporte.append({
            'producto': p.nombre,
            'stock': p.stock_actual,
            'historico_trimestral': p.total_vendido,
            'promedio_mensual': round(promedio_mensual, 1),
            'prediccion_prox_mes': round(prediccion, 0),
            'estado': estado_proyeccion
        })
    
    # Ordenar por los que más se van a vender
    reporte = sorted(reporte, key=lambda k: k['prediccion_prox_mes'], reverse=True)
    
    return render_template('reportes_predicciones.html', data=reporte, hoy=hora_peru())

# --- EN APP.PY ---

@app.route('/api/toggle_check/<int:detail_id>', methods=['POST'])
def toggle_check(detail_id):
    if 'user_id' not in session: return {'status': 'error'}, 401
    
    # Buscamos el item específico
    detalle = OrderDetail.query.get_or_404(detail_id)
    
    # Invertimos el valor (Si es True pasa a False, y viceversa)
    detalle.check_almacen = not detalle.check_almacen
    db.session.commit()
    
    return {'status': 'success', 'nuevo_estado': detalle.check_almacen}

# API SECRETA PARA CONSULTAR PRECIO EN VIVO (AJAX)
@app.route('/api/check_precio/<int:product_id>/<int:cantidad>')
def check_precio(product_id, cantidad):
    p = Product.query.get_or_404(product_id)
    
    # --- LÓGICA DE TU TESIS (Validador de 3 niveles) ---
    precio_final = 0.0
    tipo_precio = ""
    
    mitad_caja = p.unidades_por_caja / 2
    
    if cantidad >= 1 and cantidad <= 11:
        precio_final = p.precio_unidad
        tipo_precio = "Precio Unidad"
    elif cantidad >= 12 and cantidad < mitad_caja:
        precio_final = p.precio_docena
        tipo_precio = "Precio Docena"
    else:
        precio_final = p.precio_caja
        tipo_precio = "Precio Caja Mayorista"
        
    # --- SEMÁFORO DE MARGEN (OK / WARN / BLOCK) ---
    # Simulamos cálculo de margen
    margen = precio_final - p.costo_referencial
    estado = "OK"
    mensaje = "Margen saludable."
    
    if margen <= 0:
        estado = "BLOCK"
        mensaje = "ERROR: Venta con pérdida. Aumente precio."
    elif margen < (p.costo_referencial * 0.15): # Si gana menos del 15%
        estado = "WARN"
        mensaje = "ADVERTENCIA: Margen muy bajo."
        
    return {
        "precio": precio_final,
        "tipo": tipo_precio,
        "total": precio_final * cantidad,
        "estado": estado,
        "mensaje": mensaje
    }

# --- BUSCADOR DE CLIENTES (PARA SELECT2) ---
@app.route('/api/buscar_clientes_db')
def buscar_clientes_db():
    if session.get('user_id') is None: return {'results': []}
    
    q = request.args.get('q', '').strip()
    if not q: return {'results': []}
    
    # Buscar por RUC o Nombre (contiene texto)
    clientes = Client.query.filter(
        or_(
            Client.documento.ilike(f"%{q}%"),
            Client.nombre.ilike(f"%{q}%")
        )
    ).limit(10).all() # Máximo 10 resultados para ser rápido
    
    resultados = []
    for c in clientes:
        resultados.append({
            'id': c.documento, # Usamos el RUC como ID
            'text': f"{c.documento} - {c.nombre}", # Lo que se ve en la lista
            # Datos extra para llenar el formulario:
            'nombre': c.nombre,
            'direccion': c.direccion,
            'telefono': c.telefono,
            'estado': c.estado,
            'condicion': c.condicion,
            'updated': c.last_updated.strftime('%d/%m/%Y')
        })
        
    return {'results': resultados}

# --- RENOMBRAR FAMILIA (CASCADA) ---
@app.route('/categoria/editar', methods=['POST'])
def editar_categoria():
    if session.get('role') != 'admin': return {'status': 'error', 'msg': 'No autorizado'}, 403
    
    nombre_viejo = request.form.get('nombre_viejo')
    nombre_nuevo = request.form.get('nombre_nuevo', '').strip().upper()
    
    if not nombre_nuevo: return {'status': 'error', 'msg': 'Nombre vacío'}
    
    try:
        # 1. Verificar si el nuevo nombre ya existe (para evitar duplicados)
        existe = Category.query.filter_by(nombre=nombre_nuevo).first()
        if existe:
            return {'status': 'error', 'msg': f'Ya existe la familia "{nombre_nuevo}". Use la opción de borrar/fusionar manual.'}

        # 2. Actualizar Tabla Categorías
        cat = Category.query.filter_by(nombre=nombre_viejo).first()
        if cat:
            cat.nombre = nombre_nuevo
            
        # 3. Actualizar Tabla Productos (Cascada Masiva)
        # UPDATE product SET categoria = 'NUEVO' WHERE categoria = 'VIEJO'
        Product.query.filter_by(categoria=nombre_viejo).update({Product.categoria: nombre_nuevo})
        
        db.session.commit()
        return {'status': 'success', 'msg': f'Familia renombrada a {nombre_nuevo}'}
        
    except Exception as e:
        db.session.rollback()
        return {'status': 'error', 'msg': str(e)}
    
# --- EN APP.PY ---


# 1. GERENCIA HACE LA REVISIÓN INICIAL (Sin OC)
@app.route('/gestion_ventas/revision_inicial/<int:order_id>', methods=['POST'])
def revision_inicial_gerencia(order_id):
    if session.get('role') not in ['admin', 'administracion']: 
        return {'status': 'error', 'msg': 'No autorizado'}, 403

    orden = Order.query.get_or_404(order_id)
    orden.estado = 'Revision Inicial OK'
    orden.fecha_revision_inicial = hora_peru()
    
    db.session.commit()
    return {'status': 'success', 'msg': 'Revisión inicial aprobada. El vendedor ya puede pedir la OC al cliente.'}

# 2. VENDEDOR SUBE OC Y PIDE APROBACIÓN FINAL (Modificado)
@app.route('/api/confirmar_cliente/<int:order_id>', methods=['POST'])
def confirmar_cliente(order_id):
    orden = Order.query.get_or_404(order_id)
    
    if orden.vendedor_id != session.get('user_id'):
        return {'status': 'error', 'msg': 'Solo el creador puede confirmar esto.'}, 403
        
    orden.cliente_confirmado = True
    orden.fecha_confirmacion_cliente = hora_peru()
    orden.estado = 'Pendiente Aprobacion Final' 
    
    db.session.commit()
    return {'status': 'success', 'msg': 'Enviado a Gerencia para Aprobación Final (Validación OC).'}

@app.route('/gestion_ventas/aprobar/<int:order_id>', methods=['POST'])
def aprobar_cotizacion_gerencia(order_id):
    if session.get('role') not in ['admin', 'administracion']: 
        return {'status': 'error', 'msg': 'No tiene permisos de Gerencia'}, 403

    orden = Order.query.get_or_404(order_id)

    if orden.estado in ['Por Despachar', 'Despachado', 'Entregado']:
        return {'status': 'error', 'msg': 'Esta orden ya fue aprobada.'}

    try:
        errores_stock = []
        
        for detalle in orden.details:
            if detalle.product_id: 
                prod = Product.query.get(detalle.product_id)
                if prod.stock_actual < detalle.cantidad:
                    errores_stock.append(f"{prod.nombre} (Faltan {detalle.cantidad - prod.stock_actual})")
            
            if detalle.item_type == 'GLB':
                for comp in detalle.kit_components:
                    prod_c = comp.product
                    total_req = comp.cantidad_requerida * detalle.cantidad
                    if prod_c.stock_actual < total_req:
                        errores_stock.append(f"Componente {prod_c.sku} en Kit (Faltan {total_req - prod_c.stock_actual})")

        if errores_stock:
            return {'status': 'error', 'msg': 'Stock insuficiente: ' + ', '.join(errores_stock)}

        # CAMBIO: Estado más claro
        orden.estado = 'Por Despachar'
        orden.fecha_aprobacion = hora_peru() 
        orden.gerente_nombre = session.get('nombre')
        
        db.session.commit()
        
        return {'status': 'success', 'msg': f'Cotización aprobada. Pasó a Almacén como NP-{orden.id:05d} lista para despacho.'}

    except Exception as e:
        db.session.rollback()
        return {'status': 'error', 'msg': f'Error crítico: {str(e)}'}
# --- EN APP.PY (Función historial_ventas AJUSTADA) ---

@app.route('/historial_ventas')
def historial_ventas():
    # 1. Seguridad
    if 'user_id' not in session: return redirect(url_for('login'))
    
    rol = session['role']
    user_id = session['user_id']

    # Filtros fijos
    vendedores = []
    if rol in ['admin', 'administracion', 'almacen']:
        vendedores = User.query.filter_by(role='vendedor').all()
    
    # 1. QUERY BASE
    query = Order.query
    
    # --- CORRECCIÓN: LISTA EXACTA DE ESTADOS SEGÚN EL JS ---
    estados_revision = ['Por Verificar', 'Pendiente Aprobacion', 'Revision Pre-Cliente', 'Pendiente Aprobacion Final']
    
    # 2. CONTADORES (Para los globos rojos/amarillos)
    cuentas = {
        'rev': Order.query.filter(Order.estado.in_(estados_revision)).count(),
        'obs': 0
    }
    if rol == 'vendedor':
        cuentas['obs'] = Order.query.filter(Order.vendedor_id == user_id, Order.estado == 'Observado').count()
        # El vendedor solo ve en el globo amarillo sus propios pedidos que están en revisión
        cuentas['rev'] = Order.query.filter(Order.vendedor_id == user_id, Order.estado.in_(estados_revision)).count()

    # CAPTURAMOS EL NUEVO INTERRUPTOR
    solo_mias = request.args.get('solo_mias') == 'on'

    # --- 3. APLICAR LÓGICA DE PESTAÑAS Y PRIVACIDAD ---
    vista = request.args.get('vista', 'borradores')

    if vista == 'borradores':
        # AQUÍ ESTÁ EL VENDEDOR (Agregamos 'Aprobado Pre-Cliente' para que espere la OC)
        query = query.filter(Order.estado.in_([
            'Cotizacion', 'Observado', 'Stock Confirmado', 'Aprobado Pre-Cliente'
        ]))
        if rol == 'vendedor' or solo_mias: 
            query = query.filter(Order.vendedor_id == user_id)

    elif vista == 'revision':
        # AQUÍ ESTÁ GERENCIA Y ALMACÉN (Agregamos los nuevos estados de revisión)
        query = query.filter(Order.estado.in_(estados_revision))
        if rol == 'vendedor' or solo_mias: 
            query = query.filter(Order.vendedor_id == user_id)

    elif vista == 'historial':
        query = query.filter(Order.estado.in_([
            'Por Despachar', 'Entregado', 'Despachado', 'Anulado', 'Rechazado'
        ]))
        if rol == 'vendedor' or solo_mias: 
            query = query.filter(Order.vendedor_id == user_id)

    # --- 4. AHORA APLICAMOS LA BÚSQUEDA ---
    busqueda = request.args.get('busqueda')
    if busqueda:
        term_id = busqueda
        if busqueda.isdigit(): term_id = str(int(busqueda))
        query = query.join(Client).join(User).filter(
            or_(
                Client.nombre.ilike(f"%{busqueda}%"),
                Client.documento.ilike(f"%{busqueda}%"),
                User.username.ilike(f"%{busqueda}%"),
                User.nombre_completo.ilike(f"%{busqueda}%"),
                func.cast(Order.id, db.String).like(f"%{term_id}%") 
            )
        )

    # 5. OTROS FILTROS (Cliente Select, Fechas, Vendedor Dropdown)
    filtro_cliente_ruc = request.args.get('filtro_cliente')
    if filtro_cliente_ruc:
        cliente_obj_filtro = Client.query.filter_by(documento=filtro_cliente_ruc).first()
        if cliente_obj_filtro: query = query.filter(Order.cliente_id == cliente_obj_filtro.id)

    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    if fecha_inicio and fecha_fin:
        try:
            start = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            end = datetime.strptime(fecha_fin + " 23:59:59", '%Y-%m-%d %H:%M:%S')
            query = query.filter(Order.fecha.between(start, end))
        except: pass

    filtro_vendedor = request.args.get('filtro_vendedor')
    if rol != 'vendedor' and filtro_vendedor and filtro_vendedor != 'todos':
        query = query.filter(Order.vendedor_id == filtro_vendedor)

    # 6. EJECUCIÓN
    query = query.order_by(Order.fecha.desc())
    page = request.args.get('page', 1, type=int)
    per_page = 20 
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    ordenes = pagination.items 
    
    clientes = Client.query.order_by(Client.nombre).limit(2000).all()

    return render_template('historial_ventas.html', 
                           ordenes=ordenes, 
                           pagination=pagination,
                           vista_actual=vista,
                           cuentas=cuentas,
                           vendedores=vendedores,
                           clientes=clientes,
                           cliente_seleccionado=filtro_cliente_ruc,
                           hoy=hora_peru().date())

# --- RUTA PARA CARGAR LA EDICIÓN (GET) ---

@app.route('/editar_venta/<int:order_id>')
def editar_venta(order_id):
    if 'user_id' not in session: return redirect(url_for('login'))
    
    # 1. Obtener la orden
    orden = Order.query.get_or_404(order_id)
    
    # CAMBIO: Permitimos editar si está en validación, pero BLOQUEAMOS si ya se aprobó
    if orden.estado in ['Aprobado', 'Despachado', 'Entregado', 'Anulado', 'Rechazado']:
        flash('⚠️ No se puede editar un pedido que ya está aprobado o finalizado.')
        return redirect(url_for('historial_ventas'))

    # 2. Reconstruir el "Carrito" para Javascript
    items_js = []
    
    for d in orden.details:
        item = {
            'id': d.product_id,
            'sku': 'GEN', 
            'nombre': d.product.nombre if d.product else (d.nombre_personalizado or "Item"),
            'tipo': d.item_type,
            'stock': 0, 
            'cantidad': d.cantidad,
            'precio': d.precio_aplicado,
            'subtotal': d.subtotal,
            'um': getattr(d.product, 'unidad_medida', 'UND'),
            
            # --- RECUPERACIÓN A PRUEBA DE FALLOS ---
            'estado': d.product.estado if (d.product and hasattr(d.product, 'estado')) else '',
            'precioBase': getattr(d, 'precio_base', d.precio_aplicado) or d.precio_aplicado,
            'desc_tipo': getattr(d, 'desc_tipo', ''),
            'desc_valor': getattr(d, 'desc_valor', 0.0) or 0.0,
            'desc_label': getattr(d, 'desc_label', '') or ''  
        }

        # Datos específicos por tipo
        if d.item_type == 'PRODUCTO' and d.product:
            item['sku'] = d.product.sku
            item['stock'] = d.product.stock_actual
            
            # --- CORRECCIÓN DEL ERROR 'unidad_medida' ---
            # Usamos getattr para evitar que falle si la columna no existe en la BD
            item['um'] = getattr(d.product, 'unidad_medida', 'UND')
            
        elif d.item_type == 'FABRICACION':
            item['sku'] = d.product.sku if d.product else 'SRV'
            item['titulo_glb'] = d.nombre_personalizado_titulo
            item['descripcion_glb'] = d.nombre_personalizado
            item['stock'] = 9999
            item['um'] = 'SRV'

        elif d.item_type == 'GLB':
            item['sku'] = 'KIT'
            item['titulo_glb'] = d.nombre_personalizado_titulo
            item['descripcion_glb'] = d.nombre_personalizado
            item['stock'] = 9999
            item['um'] = 'GLB'
            # Reconstruir componentes
            comps = []
            for c in d.kit_components:
                comps.append({
                    'id': c.product_id,
                    'nombre': c.product.nombre,
                    'qty': c.cantidad_requerida,
                    'stock_individual': c.product.stock_actual
                })
            item['componentes'] = comps

        items_js.append(item)

    # 3. Datos Generales
    productos = Product.query.all()
    categorias = Category.query.all()
    config_tc = SystemConfig.query.get('tipo_cambio')
    
    return render_template('nueva_venta.html',
                           modo_edicion=True, 
                           orden=orden,       
                           items_json=items_js, 
                           productos=productos,
                           categorias=categorias,
                           tc=orden.tipo_cambio, 
                           updated_at=config_tc.updated_at.strftime('%d/%m') if config_tc else None)

# --- AGREGAR O CORREGIR EN APP.PY ---

@app.route('/actualizar_venta', methods=['POST'])
def actualizar_venta():
    if 'user_id' not in session: return {'status': 'error', 'msg': 'No autorizado'}, 403
    
    try:
        data = request.get_json()

        # >>> IMPRESIÓN DE CONTROL PARA ACTUALIZAR COTIZACIÓN <<<
        print("\n========================================")
        print("👉 INTENTANDO ACTUALIZAR COTIZACIÓN:")
        print(f"Agencia: {data.get('agencia')} | Calidad: {data.get('control_calidad')} | Penalidad: {data.get('penalidad')}")
        print("========================================\n")
        order_id = data.get('order_id') # Recibimos el ID
        
        if not order_id:
            return {'status': 'error', 'msg': 'No se recibió el ID de la orden para editar.'}

        # 1. BUSCAR LA ORDEN EXISTENTE
        orden = Order.query.get_or_404(order_id)
        
        # 2. VALIDAR ESTADO (Seguridad)
        # CAMBIO CLAVE: Bloqueamos SOLO los estados finales.
        if orden.estado in ['Aprobado', 'Despachado', 'Entregado', 'Anulado', 'Rechazado']:
             return {'status': 'error', 'msg': 'No se puede editar un pedido finalizado o aprobado.'}

        # 3. ACTUALIZAR DATOS CABECERA (Sobreescribir)
        orden.atencion = data.get('cliente_atte')
        orden.orden_compra = data.get('orden_compra')
        orden.condicion_pago = data.get('condicion_pago')
        orden.validez_oferta = data.get('validez_oferta')
        orden.observacion = data.get('observacion')
        
        # Fechas y Dirección
        f_entrega = data.get('fecha_entrega')
        if f_entrega:
            orden.fecha_entrega = datetime.strptime(f_entrega, '%Y-%m-%d').date()
        else:
            orden.fecha_entrega = None
            
        # Corregimos el problema de la dirección vacía en Recojo
        tipo_ent = data.get('tipo_entrega')
        dir_ent = data.get('direccion_entrega')
        if tipo_ent == 'Recojo':
            dir_ent = "RECOJO EN TIENDA / ALMACÉN"
            
        orden.tipo_entrega = tipo_ent
        orden.direccion_envio = dir_ent
        
        # --- AQUÍ GUARDAMOS LOS CAMPOS NUEVOS AL EDITAR ---
        # Aseguramos de que si llega vacío, ponga 'NO' en la base de datos
        orden.agencia = data.get('agencia') or 'NO REQUIERE'
        orden.control_calidad = data.get('control_calidad') or 'NO'
        orden.penalidad = data.get('penalidad') or 'NO'
        
        # Totales
        orden.moneda = data.get('moneda')
        orden.tipo_cambio = float(data.get('tc'))
        orden.subtotal = float(data.get('subtotal'))
        orden.igv = float(data.get('igv'))
        orden.total = float(data.get('total'))
        orden.descuento_tipo = data.get('descuento_tipo')
        orden.descuento_valor = float(data.get('descuento_valor', 0))
        orden.descuento_total = float(data.get('descuento_total', 0))

        # --- REGLA ESTRICTA DE PING-PONG ---
        # Si se edita la cotización en CUALQUIER estado permitido, TODO vuelve a cero.
        orden.estado = 'Cotizacion' 
        orden.cliente_confirmado = False
        orden.fecha_confirmacion_cliente = None
        orden.fecha_verificacion_almacen = None
        orden.almacenero_nombre = None
        orden.fecha_aprobacion = None
        orden.gerente_nombre = None
             
        # 4. ACTUALIZAR ITEMS (Estrategia: Borrar viejos y crear nuevos)
        # Esto es lo más limpio para evitar errores de duplicados en items
        OrderDetail.query.filter_by(order_id=orden.id).delete()
        
        for item in data['items']:
            tipo_item = item.get('tipo', 'PRODUCTO')
            detalle = OrderDetail(
                order_id=orden.id,
                item_type=tipo_item,
                cantidad=int(item['cantidad']),
                precio_aplicado=float(item['precio']),
                subtotal=float(item['subtotal']),
                tipo_precio_usado=item.get('tipo_precio', 'Manual'),
                    
                # --- GUARDADO DE DESCUENTOS ---
                precio_base=float(item.get('precioBase', item['precio'])),
                desc_tipo=item.get('desc_tipo', ''),
                desc_valor=float(item.get('desc_valor', 0.0)),
                desc_label=item.get('desc_label', '')
            )
            
            if tipo_item == 'PRODUCTO':
                detalle.product_id = item['id']
            elif tipo_item in ['FABRICACION', 'GLB']:
                 detalle.nombre_personalizado = item.get('descripcion_glb', item['nombre'])
                 detalle.nombre_personalizado_titulo = item.get('titulo_glb', '')
            
            db.session.add(detalle)
            db.session.flush() # Para obtener ID del detalle

            # Componentes Kit
            if tipo_item == 'GLB' and 'componentes' in item:
                for comp in item['componentes']:
                    nc = OrderKitComponent(
                        order_detail_id=detalle.id, 
                        product_id=comp['id'], 
                        cantidad_requerida=int(comp['qty'])
                    )
                    db.session.add(nc)

        db.session.commit()
        return {'status': 'success', 'order_id': orden.id, 'msg': 'Cotización actualizada correctamente.'}

    except Exception as e:
        db.session.rollback()
        return {'status': 'error', 'msg': str(e)}, 500
    

@app.route('/api/obtener_detalle_venta/<int:order_id>')
def obtener_detalle_venta(order_id):
    # 1. Seguridad básica
    if 'user_id' not in session: 
        return {'status': 'error', 'msg': 'Sesión caducada, inicie sesión.'}, 401
    
    try:
        orden = Order.query.get_or_404(order_id)
        
        # 2. Procesar Items (Productos y Kits)
        items_data = []
        for d in orden.details:
            sku_mostrado = "GEN"
            nombre_mostrado = d.nombre_personalizado or "Item sin nombre"

            # Si está vinculado a un producto real
            if d.product:
                sku_mostrado = d.product.sku
                nombre_mostrado = d.product.nombre
            # Si es Fabricación o Kit (GLB)
            elif d.item_type in ['FABRICACION', 'GLB']:
                titulo = d.nombre_personalizado_titulo or ""
                cuerpo = d.nombre_personalizado or ""
                nombre_mostrado = f"{titulo} {cuerpo}".strip()
                if d.item_type == 'FABRICACION': sku_mostrado = "SRV"
                if d.item_type == 'GLB': sku_mostrado = "KIT"

            # Lógica de Componentes (Si es un Kit)
            comps_data = []
            if d.item_type == 'GLB':
                for c in d.kit_components:
                    # Cálculo de stock visual
                    total_necesario = c.cantidad_requerida * d.cantidad
                    comps_data.append({
                        'sku': c.product.sku,
                        'nombre': c.product.nombre,
                        'cant_req': c.cantidad_requerida,
                        'cant_total': total_necesario,
                        'stock_actual': c.product.stock_actual
                    })

            items_data.append({
                'sku': sku_mostrado,
                'descripcion': nombre_mostrado,
                'estado_producto': d.product.estado if d.product else '',
                'cantidad': d.cantidad,
                'precio': d.precio_aplicado,
                'subtotal': d.subtotal,
                'tipo': d.item_type,
                'componentes': comps_data, # Enviamos la lista al HTML
                'check_almacen': d.check_almacen,
                # --- NUEVOS CAMPOS DE DESCUENTO INDIVIDUAL ---
                'precio_base': getattr(d, 'precio_base', d.precio_aplicado) or d.precio_aplicado,
                'desc_tipo': getattr(d, 'desc_tipo', ''),
                'desc_valor': getattr(d, 'desc_valor', 0.0) or 0.0,
                'desc_label': getattr(d, 'desc_label', '') or ''
            })

        # 3. Datos Generales (Manejo de nulos con "or '-'")
        data = {
            'id': orden.id,
            'vendedor_id': orden.vendedor_id,
            'fecha': orden.fecha.strftime('%d/%m/%Y %H:%M'),
            'estado': orden.estado,
            'vendedor': orden.vendedor.nombre_completo if orden.vendedor else 'Desconocido',
            
            # Cliente
            'cliente_nombre': orden.cliente.nombre if orden.cliente else 'Cliente Eliminado',
            'cliente_doc': orden.cliente.documento if orden.cliente else '-',
            'cliente_dir': orden.cliente.direccion or '-',
            'cliente_tel': orden.cliente.telefono or '-',
            
            # Reemplaza la parte de Info y Logística por esto:
            'atencion': orden.atencion or '-',
            'orden_compra': orden.orden_compra or '-',
            'archivo_oc': orden.archivo_oc,
            'condicion_pago': orden.condicion_pago or '-',
            'validez': orden.validez_oferta or '-',
            'observacion': orden.observacion or 'Ninguna',
            'motivo_rechazo': orden.motivo_rechazo or '', 
            
            # ---> DATOS NUEVOS ASEGURADOS <---
            'agencia': orden.agencia or 'NO REQUIERE',
            'control_calidad': orden.control_calidad or 'NO',
            'penalidad': orden.penalidad or 'NO',
            
            # Logística
            'tipo_entrega': orden.tipo_entrega or '-',
            'fecha_entrega': orden.fecha_entrega.strftime('%d/%m/%Y') if orden.fecha_entrega else 'A coordinar',
            'direccion_entrega': orden.direccion_envio or '-',
            
            # Financiero
            'moneda': orden.moneda,
            'subtotal': orden.subtotal,
            'igv': orden.igv,
            'total': orden.total,
            # --- NUEVOS CAMPOS DE DESCUENTO GENERAL ---
            'descuento_total': orden.descuento_total,
            'descuento_tipo': orden.descuento_tipo,
            'descuento_valor': orden.descuento_valor,
            
            'chofer_nombre': orden.chofer.username if orden.chofer else None,
            'peso_total': orden.peso_total,
            'cantidad_bultos': orden.cantidad_bultos,
            
            'items': items_data,

            # --- LÍNEA DE TIEMPO (5 PASOS) ---
            'creador': orden.vendedor.nombre_completo if orden.vendedor else 'Vendedor',
            
            'almacenero': orden.almacenero_nombre,
            'f_verif_almacen': orden.fecha_verificacion_almacen.strftime('%d/%m/%Y %H:%M') if orden.fecha_verificacion_almacen else None,
            
            # --- CORREGIDO: SE ENVÍA EL NOMBRE DEL REVISOR INICIAL ---
            'revisor_inicial': orden.revisor_inicial_nombre,
            'f_revision_inicial': orden.fecha_revision_inicial.strftime('%d/%m/%Y %H:%M') if orden.fecha_revision_inicial else None,
            
            'cliente_confirmado': orden.cliente_confirmado,
            'f_conf_cliente': orden.fecha_confirmacion_cliente.strftime('%d/%m/%Y %H:%M') if orden.fecha_confirmacion_cliente else None,
            
            'gerente': orden.gerente_nombre,
            'f_aprobacion_final': orden.fecha_aprobacion.strftime('%d/%m/%Y %H:%M') if orden.fecha_aprobacion else None,
        }
        
        return {'status': 'success', 'data': data}

    except Exception as e:
        print(f"ERROR API DETALLE: {str(e)}") # Esto saldrá en tu consola negra
        return {'status': 'error', 'msg': f"Error interno: {str(e)}"}, 500
    
# --- EN APP.PY ---

@app.route('/gestion_ventas/observar', methods=['POST'])
def observar_cotizacion():
    if session.get('role') not in ['admin', 'administracion']: 
        return {'status': 'error', 'msg': 'No tiene permisos'}, 403

    data = request.get_json()
    order_id = data.get('order_id')
    motivo = data.get('motivo')

    if not motivo:
        return {'status': 'error', 'msg': 'El motivo es obligatorio.'}

    orden = Order.query.get_or_404(order_id)
    
    orden.estado = 'Observado'
    orden.motivo_rechazo = motivo # <--- GUARDAMOS EN LA COLUMNA NUEVA
    # NO tocamos orden.observacion (ahí se queda lo que escribió el vendedor)
    
    db.session.commit()
    
    return {'status': 'success', 'msg': 'Cotización observada y devuelta.'}

    # --- NUEVA RUTA PARA REGISTRAR LA FECHA DE LA REVISIÓN INICIAL ---
@app.route('/api/aprobar_pre_cliente/<int:order_id>', methods=['POST'])
def aprobar_pre_cliente(order_id):
    if session.get('role') not in ['admin', 'administracion']: 
        return {'status': 'error', 'msg': 'No autorizado'}, 403

    orden = Order.query.get_or_404(order_id)
    orden.estado = 'Aprobado Pre-Cliente'
    orden.fecha_revision_inicial = hora_peru()
    
    # NUEVO: Guardamos el nombre del usuario que aprobó
    orden.revisor_inicial_nombre = session.get('nombre', 'Administrador')
    
    db.session.commit()
    return {'status': 'success', 'msg': 'Se ha autorizado el envío al cliente.'}

# =======================================================
# MÓDULO DE ALMACÉN: PICKING LITE (SALIDAS)
# =======================================================

@app.route('/picking_almacen')
def picking_almacen():
    if session.get('role') not in ['admin', 'almacen']: 
        return "Acceso denegado", 403
    
    # --- FILTROS PARA HISTORIAL ---
    busqueda_hist = request.args.get('busqueda_hist', '').strip()
    fecha_inicio_hist = request.args.get('fecha_inicio_hist', '')
    fecha_fin_hist = request.args.get('fecha_fin_hist', '')
    page_hist = request.args.get('page_hist', 1, type=int)

    # --- QUERIES PRINCIPALES ---
    ordenes_por_verificar = Order.query.filter_by(
        estado='Por Verificar'
    ).order_by(Order.fecha.asc()).all()
    
    ordenes_pendientes = Order.query.filter_by(
        estado='Por Despachar'
    ).order_by(Order.fecha.asc()).all()
    
    # --- HISTORIAL CON FILTROS Y PAGINACIÓN ---
    query_hist = Order.query.filter(
        Order.estado.in_(['Entregado', 'Despachado'])
    )
    
    if busqueda_hist:
        # Solución anti-errores para buscar IDs numéricos (Si buscan "00052", lo convierte a "52")
        term_id = busqueda_hist
        if busqueda_hist.isdigit(): 
            term_id = str(int(busqueda_hist))
            
        # Reemplazamos db.String por String de SQLAlchemy nativo
        query_hist = query_hist.join(Client).filter(
            or_(
                Client.nombre.ilike(f'%{busqueda_hist}%'),
                Client.documento.ilike(f'%{busqueda_hist}%'),
                func.cast(Order.id, String).ilike(f'%{term_id}%')
            )
        )
    
    if fecha_inicio_hist:
        query_hist = query_hist.filter(
            Order.fecha >= datetime.strptime(fecha_inicio_hist, '%Y-%m-%d')
        )
    if fecha_fin_hist:
        query_hist = query_hist.filter(
            Order.fecha <= datetime.strptime(fecha_fin_hist + ' 23:59:59', '%Y-%m-%d %H:%M:%S')
        )
    
    query_hist = query_hist.order_by(Order.fecha.desc())
    pagination_hist = query_hist.paginate(page=page_hist, per_page=20, error_out=False)
    ordenes_finalizadas = pagination_hist.items
    
    return render_template('picking_almacen.html', 
                           ordenes_verificar=ordenes_por_verificar,
                           ordenes=ordenes_pendientes,
                           ordenes_finalizadas=ordenes_finalizadas,
                           pagination_hist=pagination_hist,
                           busqueda_hist=busqueda_hist,
                           fecha_inicio_hist=fecha_inicio_hist,
                           fecha_fin_hist=fecha_fin_hist)


# NUEVA RUTA PARA QUE ALMACÉN APRUEBE EL STOCK FÍSICO
@app.route('/api/confirmar_stock_fisico/<int:order_id>', methods=['POST'])
def confirmar_stock_fisico(order_id):
    if session.get('role') not in ['admin', 'almacen']: 
        return {'status': 'error', 'msg': 'No autorizado'}, 403
    
    orden = Order.query.get_or_404(order_id)
    if orden.estado == 'Por Verificar':
        # CAMBIO: Pasa a 'Stock Confirmado' en lugar de Gerencia
        orden.estado = 'Stock Confirmado'
        
        orden.fecha_verificacion_almacen = hora_peru()
        orden.almacenero_nombre = session.get('nombre')
        db.session.commit()
        return {'status': 'success', 'msg': 'Stock verificado. Ahora el Vendedor debe pedir confirmación al cliente.'}
    
    return {'status': 'error', 'msg': 'El pedido no está en estado de verificación.'}

@app.route('/api/procesar_salida_almacen/<int:order_id>', methods=['POST'])
def procesar_salida_almacen(order_id):
    if session.get('role') not in ['admin', 'almacen']: 
        return {'status': 'error', 'msg': 'No autorizado'}, 403
    
    orden = Order.query.get_or_404(order_id)
    
    # CAMBIO: validar nuevo estado
    if orden.estado != 'Por Despachar':
        return {'status': 'error', 'msg': 'La cotización no está lista para despacho.'}
        
    try:
        for detalle in orden.details:
            if detalle.product_id and detalle.item_type == 'PRODUCTO':
                prod = Product.query.get(detalle.product_id)
                stock_anterior = prod.stock_actual
                prod.stock_actual -= detalle.cantidad
                
                kardex = ProductMovement(
                    product_id=prod.id,
                    user_id=session['user_id'],
                    tipo='SALIDA',
                    cantidad=detalle.cantidad,
                    stock_anterior=stock_anterior,
                    stock_nuevo=prod.stock_actual,
                    motivo=f"Salida Almacén NP-{orden.id:05d} ({orden.cliente.nombre[:15]})"
                )
                db.session.add(kardex)

            elif detalle.item_type == 'GLB':
                for comp in detalle.kit_components:
                    prod_c = comp.product
                    cantidad_total = comp.cantidad_requerida * detalle.cantidad
                    stock_ant_c = prod_c.stock_actual
                    prod_c.stock_actual -= cantidad_total
                    
                    kardex_c = ProductMovement(
                        product_id=prod_c.id,
                        user_id=session['user_id'],
                        tipo='SALIDA',
                        cantidad=cantidad_total,
                        stock_anterior=stock_ant_c,
                        stock_nuevo=prod_c.stock_actual,
                        motivo=f"Salida Kit NP-{orden.id:05d} - {detalle.nombre_personalizado_titulo[:15]}"
                    )
                    db.session.add(kardex_c)

        # CAMBIO: Estado final más claro
        orden.estado = 'Entregado'
        db.session.commit()
        
        return {'status': 'success', 'msg': 'Stock descontado. Orden cerrada como Entregada.'}
        
    except Exception as e:
        db.session.rollback()
        return {'status': 'error', 'msg': f'Error en el descargo: {str(e)}'}

# =========================================================================================
# FUNCIONES CRUD Y LOGÍSTICA PARA IMPORTBOLTS (CREAR, EDITAR, ELIMINAR, EXPORTAR, CATEGORÍAS)
# =========================================================================================

@app.route('/inventario_importbolts')
def inventario_importbolts():
    if session.get('user_id') is None: return redirect(url_for('login'))
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('busqueda', '')
    cat_filtro = request.args.get('categoria', 'todos')
    calidad_filtro = request.args.get('calidad', 'todos')
    stock_bajo = request.args.get('stock_bajo')

    # Apuntamos a la nueva tabla
    query = ProductImportBolts.query

    if search:
        query = query.filter(or_(ProductImportBolts.nombre.ilike(f"%{search}%"), ProductImportBolts.sku.ilike(f"%{search}%")))
    if cat_filtro != 'todos':
        query = query.filter(ProductImportBolts.categoria == cat_filtro)
    if calidad_filtro != 'todos':
        query = query.filter(ProductImportBolts.calidad == calidad_filtro)
    if stock_bajo == 'on':
        query = query.filter(ProductImportBolts.stock_actual <= ProductImportBolts.stock_minimo)

    cats_db = CategoryImportBolts.query.order_by(CategoryImportBolts.nombre).all()
    lista_categorias = [c.nombre for c in cats_db]
    calidades = db.session.query(ProductImportBolts.calidad).distinct().order_by(ProductImportBolts.calidad).all()
    lista_calidades = [c[0] for c in calidades if c[0]]
    estados = db.session.query(ProductImportBolts.estado).filter(ProductImportBolts.estado != None, ProductImportBolts.estado != '').distinct().order_by(ProductImportBolts.estado).all()
    lista_estados = [e[0] for e in estados]

    if stock_bajo == 'on':
        query = query.order_by(ProductImportBolts.stock_actual.asc())
    else:
        query = query.order_by(ProductImportBolts.id.asc())

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    productos = pagination.items
    
    info_importacion = SystemConfig.query.get('ultima_importacion_importbolts')
    
    return render_template('inventario_importbolts.html', 
                           productos=productos, 
                           lista_categorias=lista_categorias, 
                           lista_calidades=lista_calidades,
                           pagination=pagination,
                           search=search,
                           cat_filtro=cat_filtro,
                           calidad_filtro=calidad_filtro,
                           stock_bajo=stock_bajo, 
                           limit=per_page,
                           lista_estados=lista_estados,
                           info_importacion=info_importacion)
    

@app.route('/producto_importbolts/ajustar_stock', methods=['POST'])
def ajustar_stock_importbolts():
    if session.get('role') not in ['admin', 'almacen']: return "No autorizado", 403
    
    prod_id = request.form['prod_id']
    tipo_ajuste = request.form['tipo']
    cantidad = int(request.form['cantidad'])
    motivo_texto = request.form['motivo']
    url_origen = request.form.get('url_origen')

    prod = ProductImportBolts.query.get(prod_id)
    stock_antes = prod.stock_actual
    
    tipo_kardex = ""
    
    if tipo_ajuste == 'ingreso':
        prod.stock_actual += cantidad
        tipo_kardex = "ENTRADA"
        flash(f'Ingreso registrado: +{cantidad} en {prod.sku}')
    else:
        prod.stock_actual -= cantidad
        tipo_kardex = "SALIDA"
        flash(f'Salida registrada: -{cantidad} en {prod.sku}')
        
    kardex = ProductMovementImportBolts(
        product_id=prod.id,
        user_id=session['user_id'],
        tipo=tipo_kardex,
        cantidad=cantidad,
        stock_anterior=stock_antes,
        stock_nuevo=prod.stock_actual,
        motivo=motivo_texto
    )
    db.session.add(kardex)
    db.session.commit()
    
    if url_origen:
        return redirect(url_origen)
    
    return redirect(url_for('inventario_importbolts'))

@app.route('/producto_importbolts/importar', methods=['POST'])
def importar_excel_importbolts():
    import gc
    import traceback
    from openpyxl import load_workbook

    if session.get('role') not in ['admin', 'almacen']:
        return "No autorizado", 403

    if 'archivo_excel' not in request.files:
        flash('No se seleccionó ningún archivo')
        return redirect(url_for('inventario_importbolts'))

    archivo = request.files['archivo_excel']
    if not archivo or archivo.filename == '':
        return redirect(url_for('inventario_importbolts'))

    filename = secure_filename(archivo.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    archivo.save(filepath)

    nuevos = 0
    actualizados = 0

    try:
        wb = load_workbook(filename=filepath, read_only=True, data_only=True)
        ws = wb['STOCK'] if 'STOCK' in wb.sheetnames else wb.active
        
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip().upper() if h is not None else '' for h in header_row]

        def get_col(row_vals, *posibles_nombres):
            for nombre in posibles_nombres:
                if nombre in headers:
                    idx = headers.index(nombre)
                    if idx < len(row_vals): return row_vals[idx]
            return None

        hora_actual = hora_peru()
        usuario_actual = session.get('username', 'Sistema')
        user_id_actual = session.get('user_id')

        # Cache de la tabla nueva
        skus_existentes = {p.sku: p.id for p in db.session.query(ProductImportBolts.id, ProductImportBolts.sku).all()}
        cats_existentes = {c.nombre: c.prefijo for c in db.session.query(CategoryImportBolts.nombre, CategoryImportBolts.prefijo).all()}
        db.session.expunge_all()
        gc.collect()

        batch_updates = []
        batch_inserts = []
        batch_kardex  = []
        BATCH_SIZE = 100

        for row_vals in ws.iter_rows(min_row=2, values_only=True):
            sku_raw = get_col(row_vals, 'CÓDIGO', 'CODIGO', 'SKU', 'CÓDIGO ')
            if sku_raw is None: continue
            sku = str(sku_raw).strip()
            if sku.endswith('.0'): sku = sku[:-2]
            if not sku or sku.lower() in ('nan', 'none', ''): continue

            # Usar tus mismas funciones de limpieza `clean_str` y `limpiar_campo` aquí (las copias de tu código original)
            def clean_str(val, default=''):
                if val is None: return default
                s = str(val).strip()
                return default if s.lower() in ('nan', 'none', '') else s
            def clean_int(val, default=0):
                try:
                    v = str(val).replace(',', '').strip()
                    return int(float(v)) if v and v.lower() not in ('nan', 'none', '') else default
                except: return default
            def clean_float(val, default=0.0):
                try:
                    v = str(val).replace(',', '').strip()
                    return float(v) if v and v.lower() not in ('nan', 'none', '') else default
                except: return default

            import unicodedata
            def limpiar_campo(valor, max_len):
                if not valor: return ''
                limpio = ''.join(c for c in str(valor) if unicodedata.category(c)[0] not in ('C',) and ord(c) < 65536).strip()
                return limpio[:max_len]

            nombre    = limpiar_campo(clean_str(get_col(row_vals, 'DESCRIPCIÓN', 'DESCRIPCION', 'NOMBRE'), 'Sin Nombre'), 490)
            familia   = limpiar_campo(clean_str(get_col(row_vals, 'FAMILIA', 'CATEGORIA'), 'GENERAL'), 190)
            calidad   = limpiar_campo(clean_str(get_col(row_vals, 'CALIDAD'), '-'), 190)
            ubicacion = limpiar_campo(clean_str(get_col(row_vals, 'UBICACION', 'UBICACIÓN')), 190)
            estado_v  = limpiar_campo(clean_str(get_col(row_vals, 'ESTADO')).upper(), 90)
            if estado_v == 'OK': estado_v = ''
            
            stock_val = clean_int(get_col(row_vals, 'CANT. ACT.', 'STOCK', 'CANTIDAD', 'CANT.ACT.'))
            min_val   = clean_int(get_col(row_vals, 'STOCK MÍNIMO', 'STOCK MINIMO', 'MINIMO'), 10)
            precio_unit = clean_float(get_col(row_vals, 'PRECIO UNI.','PRECIO UNIT', 'PRECIO UNIDAD', 'P. UNIT', 'PRECIO_UNIT', 'PRECIO UNITARIO'))

            if familia not in cats_existentes:
                base = "".join(c for c in familia[:3].upper() if c.isalnum()) or "GEN"
                prefijo_final = base
                n = 1
                prefijos_usados = set(cats_existentes.values())
                while prefijo_final in prefijos_usados:
                    prefijo_final = f"{base[:2]}{n}"
                    n += 1
                nuevo_cat = CategoryImportBolts(nombre=familia, prefijo=prefijo_final, contador=0)
                db.session.add(nuevo_cat)
                db.session.flush()
                cats_existentes[familia] = prefijo_final

            if sku in skus_existentes:
                upd = {
                    'sku': sku, 'nombre': nombre, 'categoria': familia, 'calidad': calidad, 
                    'ubicacion': ubicacion, 'estado': estado_v, 'stock_actual': stock_val, 
                    'stock_minimo': min_val, 'fecha_actualizacion': hora_actual, 'actualizado_por': usuario_actual
                }
                if precio_unit > 0:
                    upd['tiene_precio'] = True
                    upd['precio_unidad'] = precio_unit
                else:
                    upd['tiene_precio'] = False
                    upd['precio_unidad'] = 0.0
                batch_updates.append(upd)
                actualizados += 1
            else:
                nuevo_prod = ProductImportBolts(
                    sku=sku, nombre=nombre, categoria=familia, calidad=calidad, ubicacion=ubicacion, 
                    stock_actual=stock_val, stock_minimo=min_val, precio_unidad=precio_unit, 
                    precio_caja=0.0, precio_docena=precio_unit, costo_referencial=0.0, estado=estado_v,
                    fecha_actualizacion=hora_actual, actualizado_por=usuario_actual
                )
                batch_inserts.append(nuevo_prod)
                skus_existentes[sku] = -1
                nuevos += 1

            if (nuevos + actualizados) % BATCH_SIZE == 0:
                for upd in batch_updates:
                    if upd.get('tiene_precio'):
                        db.session.execute(text("""
                            UPDATE product_importbolts SET
                                nombre=:nombre, categoria=:categoria, calidad=:calidad, ubicacion=:ubicacion, estado=:estado,
                                stock_actual=:stock_actual, stock_minimo=:stock_minimo, precio_unidad=:precio_unidad,
                                precio_docena=:precio_unidad, fecha_actualizacion=:fecha_actualizacion, actualizado_por=:actualizado_por
                            WHERE sku=:sku
                        """), upd)
                    else:
                        db.session.execute(text("""
                            UPDATE product_importbolts SET
                                nombre=:nombre, categoria=:categoria, calidad=:calidad, ubicacion=:ubicacion, estado=:estado,
                                stock_actual=:stock_actual, stock_minimo=:stock_minimo, fecha_actualizacion=:fecha_actualizacion, actualizado_por=:actualizado_por
                            WHERE sku=:sku
                        """), upd)
                
                if batch_inserts:
                    db.session.add_all(batch_inserts)
                    db.session.flush()
                    for p in batch_inserts:
                        if p.stock_actual > 0 and p.id:
                            batch_kardex.append(ProductMovementImportBolts(
                                product_id=p.id, user_id=user_id_actual, tipo='ENTRADA', cantidad=p.stock_actual,
                                stock_anterior=0, stock_nuevo=p.stock_actual, motivo="Saldo Inicial (Importación)"
                            ))
                    if batch_kardex: db.session.add_all(batch_kardex)

                db.session.commit()
                batch_updates = []; batch_inserts = []; batch_kardex = []

        # ÚLTIMO BATCH (Residuos)
        for upd in batch_updates:
            if upd.get('tiene_precio'):
                db.session.execute(text("""
                    UPDATE product_importbolts SET
                        nombre=:nombre, categoria=:categoria, calidad=:calidad, ubicacion=:ubicacion, estado=:estado,
                        stock_actual=:stock_actual, stock_minimo=:stock_minimo, precio_unidad=:precio_unidad,
                        precio_docena=:precio_unidad, fecha_actualizacion=:fecha_actualizacion, actualizado_por=:actualizado_por
                    WHERE sku=:sku
                """), upd)
            else:
                db.session.execute(text("""
                    UPDATE product_importbolts SET
                        nombre=:nombre, categoria=:categoria, calidad=:calidad, ubicacion=:ubicacion, estado=:estado,
                        stock_actual=:stock_actual, stock_minimo=:stock_minimo, fecha_actualizacion=:fecha_actualizacion, actualizado_por=:actualizado_por
                    WHERE sku=:sku
                """), upd)
                
        if batch_inserts:
            db.session.add_all(batch_inserts)
            db.session.flush()
            for p in batch_inserts:
                if p.stock_actual > 0 and p.id:
                    batch_kardex.append(ProductMovementImportBolts(
                        product_id=p.id, user_id=user_id_actual, tipo='ENTRADA', cantidad=p.stock_actual,
                        stock_anterior=0, stock_nuevo=p.stock_actual, motivo="Saldo Inicial (Importación)"
                    ))
            if batch_kardex: db.session.add_all(batch_kardex)

        db.session.commit()
        
        # Registro en SystemConfig separado
        config_import = SystemConfig.query.get('ultima_importacion_importbolts')
        hora_final = hora_peru()
        if not config_import:
            config_import = SystemConfig(key='ultima_importacion_importbolts', value='EXITOSO', updated_at=hora_final, updated_by=usuario_actual)
            db.session.add(config_import)
        else:
            config_import.updated_at = hora_final
            config_import.updated_by = usuario_actual
        db.session.commit()

        flash(f'✅ Importación completada en ImportBolts: {nuevos} nuevos, {actualizados} actualizados.')

    except Exception as e:
        db.session.rollback()
        print(f"ERROR IMPORTACIÓN IMPORTBOLTS:\n{traceback.format_exc()}")
        flash(f'Error en la importación: {str(e)}')
    finally:
            try:
                if os.path.exists(filepath): 
                    os.remove(filepath)
            except Exception as err:
                print(f"Aviso: Windows bloqueó la eliminación del temporal, se ignorará. Error: {err}")
            gc.collect()

    return redirect(url_for('inventario_importbolts'))




@app.route('/producto_importbolts/nuevo', methods=['POST'])
def nuevo_producto_importbolts():
    if session.get('role') not in ['admin', 'almacen']: 
        return {'status': 'error', 'msg': 'No autorizado'}, 403
    
    try:
        sku_manual = request.form.get('sku', '').strip()
        familia_nombre = request.form['categoria_nombre']
        nombre = request.form['nombre'].strip()
        calidad = request.form['calidad'].strip()
        ubicacion = request.form.get('ubicacion', '').strip()
        estado_val = request.form.get('estado', '').strip().upper()
        if estado_val == 'OK': estado_val = ""
        
        try:
            stock = int(request.form['stock'])
            stock_min = int(request.form.get('stock_minimo', 10))
            p_unidad = float(request.form['p_unidad']) if request.form['p_unidad'] else 0.0
            p_caja = float(request.form['p_caja']) if request.form['p_caja'] else 0.0
        except:
            return {'status': 'error', 'msg': 'Formato numérico inválido'}

        if not nombre: return {'status': 'error', 'msg': 'Falta la descripción'}
        if stock < 0 or p_unidad < 0 or stock_min < 0: return {'status': 'error', 'msg': 'No negativos'}

        sku_final = ""
        if sku_manual:
            sku_final = sku_manual.upper()
            if ProductImportBolts.query.filter_by(sku=sku_final).first():
                return {'status': 'error', 'msg': f'El SKU "{sku_final}" ya existe en ImportBolts.'}
        else:
            cat = CategoryImportBolts.query.filter_by(nombre=familia_nombre).first()
            if not cat:
                base = "".join(c for c in familia_nombre[:3].upper() if c.isalnum()) or "GEN"
                cat = CategoryImportBolts(nombre=familia_nombre, prefijo=base, contador=0)
                db.session.add(cat)
            cat.contador += 1
            sku_final = f"{cat.prefijo}-{str(cat.contador).zfill(4)}"

        nuevo = ProductImportBolts(
            sku=sku_final, nombre=nombre, categoria=familia_nombre, calidad=calidad,
            ubicacion=ubicacion, stock_actual=stock, stock_minimo=stock_min,
            precio_unidad=p_unidad, precio_caja=p_caja, precio_docena=p_unidad * 0.9, costo_referencial=0.0
        )
        db.session.add(nuevo)
        db.session.flush()
        
        if stock > 0:
            kardex = ProductMovementImportBolts(
                product_id=nuevo.id, user_id=session['user_id'], tipo='ENTRADA',
                cantidad=stock, stock_anterior=0, stock_nuevo=stock, motivo="Saldo Inicial (ImportBolts)"
            )
            db.session.add(kardex)

        registrar_log(f"Creó producto ImportBolts {sku_final}", "bi-plus-circle-fill", "text-success")
        db.session.commit()
        
        return {'status': 'success', 'msg': 'Creado', 'sku': sku_final}
        
    except Exception as e:
        db.session.rollback()
        return {'status': 'error', 'msg': str(e)}

@app.route('/producto_importbolts/editar', methods=['POST'])
def editar_producto_importbolts():
    if session.get('role') != 'admin': return "Acceso denegado", 403
    
    try:
        prod_id = request.form['prod_id']
        url_origen = request.form.get('url_origen')
        
        prod = ProductImportBolts.query.get(prod_id)
        if not prod:
            flash('Producto no encontrado')
            return redirect(url_for('inventario_importbolts'))

        nombre = request.form['nombre'].strip()
        nueva_familia = request.form.get('categoria', '').strip()
        nueva_calidad = request.form.get('calidad', '').strip()
        estado_val = request.form.get('estado', '').strip().upper()
        if estado_val == 'OK': estado_val = ""
        
        if not nombre or not nueva_familia or not nueva_calidad:
            flash('⛔ Error: Faltan datos obligatorios.')
            return redirect(url_for('inventario_importbolts'))

        prod.nombre = nombre
        prod.stock_minimo = int(request.form.get('stock_minimo', 10))
        prod.precio_unidad = float(request.form['p_unidad'])
        prod.precio_caja = float(request.form['p_caja'])
        prod.ubicacion = request.form.get('ubicacion', '').strip()
        prod.categoria = nueva_familia
        prod.calidad = nueva_calidad
        prod.estado = estado_val
        
        registrar_log(f"Editó producto ImportBolts {prod.sku}", "bi-pencil-fill", "text-warning")
        db.session.commit()
        flash('✅ Producto actualizado correctamente.')
        
        if url_origen: return redirect(url_origen)
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error al editar: {str(e)}')
        
    return redirect(url_for('inventario_importbolts'))

@app.route('/producto_importbolts/eliminar/<int:prod_id>')
def eliminar_producto_importbolts(prod_id):
    if session.get('role') != 'admin': 
        flash('No tiene permisos para eliminar.')
        return redirect(url_for('inventario_importbolts'))
    
    try:
        prod = ProductImportBolts.query.get_or_404(prod_id)
        sku_eliminado = prod.sku

        # Limpiar Kardex de ImportBolts
        ProductMovementImportBolts.query.filter_by(product_id=prod_id).delete()
        db.session.delete(prod)
        db.session.commit()
        
        flash(f'✅ Producto {sku_eliminado} eliminado de ImportBolts.')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar: {str(e)}')

    return redirect(request.referrer or url_for('inventario_importbolts'))

@app.route('/producto_importbolts/exportar')
def exportar_excel_importbolts():
    import gc
    import io
    import pandas as pd
    if session.get('role') not in ['admin', 'almacen', 'administracion']: return "No autorizado", 403
    
    productos = db.session.query(
        ProductImportBolts.sku, ProductImportBolts.nombre, ProductImportBolts.categoria, 
        ProductImportBolts.calidad, ProductImportBolts.ubicacion, ProductImportBolts.stock_actual, 
        ProductImportBolts.stock_minimo, ProductImportBolts.precio_unidad, ProductImportBolts.precio_caja
    ).all()
    
    data = []
    for p in productos:
        data.append({
            'CÓDIGO': p.sku, 'DESCRIPCIÓN': p.nombre, 'FAMILIA': p.categoria, 'CALIDAD': p.calidad,
            'UBICACION': p.ubicacion, 'STOCK ACTUAL': p.stock_actual, 'STOCK MÍNIMO': p.stock_minimo,
            'PRECIO UNIT': p.precio_unidad, 'PRECIO CAJA': p.precio_caja
        })
    
    del productos
    db.session.expunge_all()
    gc.collect()
    
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Inventario')
        worksheet = writer.sheets['Inventario']
        for idx, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
            worksheet.set_column(idx, idx, max_len)

    output.seek(0)
    del data; del df; gc.collect()
    
    return send_file(
        output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=f'Inventario_ImportBolts_Data_{hora_peru().strftime("%Y%m%d")}.xlsx'
    )

@app.route('/categoria_importbolts/nueva', methods=['POST'])
def nueva_categoria_importbolts():
    if session.get('role') not in ['admin', 'almacen']: return "No autorizado", 403
    
    nombre = request.form.get('cat_nombre', '').strip().upper()
    prefijo = request.form.get('cat_prefijo', '').strip().upper()
    
    if not nombre or not prefijo:
        flash('Error: Nombre y Prefijo son obligatorios')
        return redirect(url_for('inventario_importbolts'))
        
    if CategoryImportBolts.query.filter_by(nombre=nombre).first():
        flash('Error: Esa familia ya existe en ImportBolts.')
        return redirect(url_for('inventario_importbolts'))
        
    try:
        nueva = CategoryImportBolts(nombre=nombre, prefijo=prefijo, contador=0)
        db.session.add(nueva)
        db.session.commit()
        flash(f'✅ Familia "{nombre}" creada en ImportBolts.')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}')
        
    return redirect(url_for('inventario_importbolts'))

@app.route('/api/calidades_de_familia_importbolts', methods=['POST'])
def calidades_de_familia_importbolts():
    if session.get('user_id') is None: return {'status': 'error'}, 403
    familia = request.form.get('familia')
    try:
        calidades = db.session.query(ProductImportBolts.calidad)\
            .filter_by(categoria=familia).distinct()\
            .order_by(ProductImportBolts.calidad).all()
        return {'status': 'success', 'calidades': [c[0] for c in calidades if c[0]]}
    except Exception as e:
        return {'status': 'error', 'msg': str(e)}


@app.route('/categoria_importbolts/eliminar', methods=['POST'])
def eliminar_categoria_importbolts():
    if session.get('role') != 'admin':
        return {'status': 'error', 'msg': 'No autorizado'}, 403
    cat_nombre = request.form.get('nombre_cat')
    if ProductImportBolts.query.filter_by(categoria=cat_nombre).count() > 0:
        return {'status': 'error', 'msg': f'⛔ "{cat_nombre}" tiene productos asociados.'}
    cat = CategoryImportBolts.query.filter_by(nombre=cat_nombre).first()
    if cat:
        db.session.delete(cat)
        db.session.commit()
        return {'status': 'success', 'msg': f'Familia "{cat_nombre}" eliminada.'}
    return {'status': 'error', 'msg': 'La familia no existe.'}


@app.route('/categoria_importbolts/editar', methods=['POST'])
def editar_categoria_importbolts():
    if session.get('role') != 'admin':
        return {'status': 'error', 'msg': 'No autorizado'}, 403
    nombre_viejo = request.form.get('nombre_viejo')
    nombre_nuevo = request.form.get('nombre_nuevo', '').strip().upper()
    if not nombre_nuevo: return {'status': 'error', 'msg': 'Nombre vacío'}
    if CategoryImportBolts.query.filter_by(nombre=nombre_nuevo).first():
        return {'status': 'error', 'msg': f'Ya existe la familia "{nombre_nuevo}".'}
    try:
        cat = CategoryImportBolts.query.filter_by(nombre=nombre_viejo).first()
        if cat: cat.nombre = nombre_nuevo
        ProductImportBolts.query.filter_by(categoria=nombre_viejo)\
            .update({ProductImportBolts.categoria: nombre_nuevo})
        db.session.commit()
        return {'status': 'success', 'msg': f'Familia renombrada a {nombre_nuevo}'}
    except Exception as e:
        db.session.rollback()
        return {'status': 'error', 'msg': str(e)}


@app.route('/api/preview_minimos_importbolts', methods=['POST'])
def preview_minimos_importbolts():
    if session.get('role') != 'admin': return {'status': 'error', 'msg': 'No autorizado'}, 403
    familia = request.form.get('categoria_nombre')
    calidad = request.form.get('calidad_nombre')
    query = ProductImportBolts.query.filter_by(categoria=familia)
    if calidad and calidad != 'TODAS':
        query = query.filter_by(calidad=calidad)
    productos = query.order_by(ProductImportBolts.sku.asc()).all()
    lista = [{'id': p.id, 'sku': p.sku, 'nombre': p.nombre, 'min_actual': p.stock_minimo} for p in productos]
    return {'status': 'success', 'total': len(lista), 'productos': lista}


@app.route('/config/minimos_masivos_importbolts', methods=['POST'])
def actualizar_minimos_masivos_importbolts():
    if session.get('role') not in ['admin', 'almacen']:
        return {'status': 'error', 'msg': 'No autorizado'}, 403
    data = request.get_json()
    if not data: return {'status': 'error', 'msg': 'No se recibieron datos.'}
    ids = data.get('ids', [])
    try:
        nuevo_minimo = int(data.get('nuevo_minimo'))
        if nuevo_minimo < 0: raise ValueError()
    except:
        return {'status': 'error', 'msg': 'Cantidad inválida.'}
    if not ids: return {'status': 'error', 'msg': 'Seleccione al menos un producto.'}
    try:
        resultado = ProductImportBolts.query.filter(ProductImportBolts.id.in_(ids)).update(
            {ProductImportBolts.stock_minimo: nuevo_minimo}, synchronize_session=False)
        db.session.commit()
        return {'status': 'success', 'msg': f'Se actualizaron {resultado} productos.'}
    except Exception as e:
        db.session.rollback()
        return {'status': 'error', 'msg': str(e)}

# --- RUTA SECRETA PARA INICIALIZAR LA BASE DE DATOS EN RENDER ---
# --- RUTA SECRETA PARA CREAR/RESETEAR LA BASE DE DATOS DESDE EL NAVEGADOR ---

@app.route('/fix_estados_y_fechas')
def fix_estados_y_fechas():
    try:
        with db.engine.connect() as conn:
            # En SQLite no se puede modificar el tipo de columna, pero no importa 
            # porque SQLite no aplica el límite estricto de caracteres.
            # Solo agregamos la columna que falta:
            conn.execute(text('ALTER TABLE "order" ADD COLUMN fecha_revision_inicial TIMESTAMP'))
            conn.commit()
        return "<h2>✅ Base de datos actualizada: Columna fecha_revision_inicial agregada con éxito.</h2>"
    except Exception as e:
        return f"<h2>Aviso: {str(e)} (Si dice 'duplicate column name', significa que ya se agregó).</h2>"
    
@app.route('/fix_revisor')
def fix_revisor():
    try:
        from sqlalchemy import text
        with db.engine.connect() as conn:
            # Agregamos la nueva columna a la base de datos
            conn.execute(text('ALTER TABLE "order" ADD COLUMN revisor_inicial_nombre VARCHAR(100)'))
            conn.commit()
        return "<h2>✅ Base de datos actualizada: Nombre de revisor agregado.</h2>"
    except Exception as e:
        return f"<h2>Aviso: {str(e)}</h2>"

@app.route('/reset_total_db_secreto')
def reset_total_db_secreto():
    try:
        # Importamos aquí para evitar errores de referencia circular
        from models import User
        from werkzeug.security import generate_password_hash
        
        # 1. Borrar todo y crear de nuevo las tablas
        db.drop_all()
        db.create_all()
        
        # 2. Crear los usuarios base de tu sistema
        admin = User(
            username='admin', 
            password=generate_password_hash('181404'),  # Contraseña inicial
            nombre_completo='Administrador General',
            role='admin'
        )
        
        db.session.add_all([admin])
        db.session.commit()
        
        return "<h1>¡Éxito!</h1><p>La base de datos ha sido reiniciada, las tablas han sido creadas y los usuarios base están listos. Ya puedes iniciar sesión.</p>"
        
    except Exception as e:
        db.session.rollback()
        return f"<h1>Error al configurar la Base de Datos:</h1><p>{str(e)}</p>"

@app.route('/fix_render_db')
def fix_render_db():
    try:
        from sqlalchemy import text
        with db.engine.connect() as conn:
            # 1. Ampliar el límite de caracteres de la columna 'estado' (VITAL)
            try:
                conn.execute(text('ALTER TABLE "order" ALTER COLUMN estado TYPE VARCHAR(50)'))
            except Exception as e:
                print(f"Aviso estado: {e}")

            # 2. Agregar fecha de revisión inicial
            try:
                conn.execute(text('ALTER TABLE "order" ADD COLUMN fecha_revision_inicial TIMESTAMP'))
            except Exception as e:
                print(f"Aviso fecha_rev: {e}")

            # 3. Agregar nombre del revisor
            try:
                conn.execute(text('ALTER TABLE "order" ADD COLUMN revisor_inicial_nombre VARCHAR(100)'))
            except Exception as e:
                print(f"Aviso revisor: {e}")

            conn.commit()
        return "<h2>✅ Base de datos en Render actualizada. Las columnas y tamaños ya coinciden con tu entorno local.</h2>"
    except Exception as e:
        return f"<h2>❌ Error general: {str(e)}</h2>"

# --- ARRANQUE DE LA APLICACIÓN ---
if __name__ == '__main__':
    # host='0.0.0.0' permite que otras PCs/celulares en la red te vean
    app.run(debug=True, host='0.0.0.0', port=5000)
